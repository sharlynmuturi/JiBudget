# Create your views here.
from django.shortcuts import render, redirect, get_object_or_404
from .models import Income, Expense, SavingsGoal, ExpenseCategory, ExpenseSubCategory, BudgetItem, IncomeCategory, IncomeSubCategory, ExpenseSource, WalletTransfer
from .forms import ExpenseForm, SavingsGoalForm, IncomeForm, BudgetItemForm, ExpenseCategoryForm, ContactForm, CustomUserCreationForm, WalletTransferForm
from django.db import models
from django.db.models import Sum, Q
from django.db.models.functions import TruncMonth
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, EmailMessage
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import date, datetime, timedelta
from calendar import monthrange
from decimal import Decimal, InvalidOperation
from decimal import Decimal, getcontext
import calendar
import openpyxl
import json
from collections import defaultdict, Counter
from django.contrib.auth.decorators import login_required
from operator import itemgetter
from collections import defaultdict, Counter
from django.contrib.auth.decorators import login_required
from operator import itemgetter
import json
from django.db.models import Min, Max
from collections import defaultdict, Counter
from operator import itemgetter
from django.db.models import Min, Max
import json
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from collections import defaultdict
from datetime import date, timedelta
from calendar import monthrange
from decimal import Decimal
import calendar, json
from django.shortcuts import render
from collections import defaultdict
import json
from datetime import datetime, date
import calendar
from datetime import datetime, date
import calendar
import json
from collections import defaultdict
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model, logout
from django.shortcuts import redirect
from django.contrib import messages
# views.py
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import LoanAccount
import json
from django.core.serializers.json import DjangoJSONEncoder
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from decimal import Decimal
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.serializers.json import DjangoJSONEncoder
import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from .models import LoanAccount, LoanPayment, ExpenseCategory, ExpenseSubCategory, Expense
from .forms import LoanAccountForm, LoanPaymentForm
from .utils import loan_amortization_schedule
import datetime

def home(request):
    return render(request, 'budget/home.html')

def about(request):
    return render(request, 'budget/about.html')

def signup_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.email = form.cleaned_data['email']
            user.save()
            login(request, user)  # Automatically log in the user
            return redirect('onboarding_view')
    else:
        form = CustomUserCreationForm()
    return render(request, 'budget/signup.html', {'form': form})



def contact_view(request):
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            email = form.cleaned_data['email']
            message = form.cleaned_data['message']

            email_subject = f"New Contact Message from {name}"
            email_body = f"""
You have received a new message from your website contact form.

Name: {name}
Email: {email}

Message:
{message}
            """

            try:
                email_message = EmailMessage(
                    subject=email_subject,
                    body=email_body,
                    from_email='sharlynnmuturi@gmail.com',  # Must match EMAIL_HOST_USER
                    to=['sharlynnmuturi@gmail.com'],         # Your inbox
                    reply_to=[email],                        # Lets you reply to sender
                )
                email_message.send()
                messages.success(request, 'Thanks for reaching out! We have received your message and will get back to you soon.')
                return redirect('contact')
            except Exception as e:
                messages.error(request, f'Something went wrong: {e}')
    else:
        form = ContactForm()

    return render(request, 'budget/contact.html', {'form': form})


@login_required
def my_dashboard(request):
    user = request.user
    today = date.today()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = first_day_of_month
            end_date = last_day_of_month
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    # Full data for other dashboard features
    incomes = Income.objects.filter(
        user=user,
        date__range=[start_date, end_date],
        amount__gt=0
    ).order_by('-date')
    
    expenses = Expense.objects.filter(
        user=user,
        date__range=[start_date, end_date],
        amount__gt=0
    ).order_by('-date')



    for inc in incomes:
        inc.is_upcoming = inc.date > today

    for exp in expenses:
        exp.is_planned = exp.date > today



    # Planned incomes and expenses (after today)
    planned_incomes_qs = Income.objects.filter(
        user=user,
        date__gt=today,
        date__range=[start_date, end_date]
    )
    planned_expenses_qs = Expense.objects.filter(
        user=user,
        date__gt=today,
        date__range=[start_date, end_date]
    )
    
    # Totals
    planned_income_total = planned_incomes_qs.aggregate(total=Sum('amount'))['total'] or 0
    planned_expense_total = planned_expenses_qs.aggregate(total=Sum('amount'))['total'] or 0


    # Incomes and expenses up to today (in range)
    actual_incomes_qs = Income.objects.filter(
        user=user,
        date__lte=today,
        date__range=[start_date, end_date]
    )
    actual_expenses_qs = Expense.objects.filter(
        user=user,
        date__lte=today,
        date__range=[start_date, end_date]
    )

    # Totals
    actual_income_total = actual_incomes_qs.aggregate(total=Sum('amount'))['total'] or 0
    actual_expense_total = actual_expenses_qs.aggregate(total=Sum('amount'))['total'] or 0
    actual_balance = actual_income_total - actual_expense_total

    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)
    balance = total_income - total_expense

    # expenses_for_budget_vs_actual = expenses
    expenses_for_budget_vs_actual = Expense.objects.filter(
        user=request.user, date__range=[start_date, end_date]
    ).exclude(category__isnull=True, subcategory__isnull=True)

    budget_items = BudgetItem.objects.filter(user=user, month__range=[start_date, end_date])

    summary = defaultdict(lambda: {'budgeted': 0, 'actual': 0})
    total_budget = sum(item.forecasted_amount for item in budget_items)
    total_spent = sum(exp.amount for exp in expenses_for_budget_vs_actual)
    total_combined = total_budget + total_spent
    budget_pct = 100
    spent_pct = round((total_spent / total_budget * 100), 1) if total_budget > 0 else 0

    # Future budgets
    future_budget_items = BudgetItem.objects.filter(user=user, month__gt=end_date)
    future_budget = sum(item.forecasted_amount for item in future_budget_items)
    all_budget = total_budget + future_budget

    income_minus_budget = total_income - total_budget
    unassigned_budget = max(total_income - total_budget, 0)

    current_month_start = date(today.year, today.month, 1)
    past_incomes = Income.objects.filter(user=user, date__lt=current_month_start)
    past_expenses = Expense.objects.filter(user=user, date__lt=current_month_start)
    past_income_total = sum(i.amount for i in past_incomes)
    past_expense_total = sum(e.amount for e in past_expenses)
    past_balance = past_income_total - past_expense_total

    cash_on_hand = past_balance + total_income
    
    actual_cash_on_hand = past_balance + actual_income_total
    
    # Sum all saved amounts across user's savings jars
    total_saved_so_far = SavingsGoal.objects.filter(user=user).aggregate(
        total=Sum('saved_so_far')
    )['total'] or 0
    
    # Adjusted available_to_budget by subtracting total saved so far
    # available_to_budget = max(cash_on_hand - all_budget - total_saved_so_far, 0)
    
    #actual_available_to_budget = max(actual_cash_on_hand - all_budget - total_saved_so_far, 0)
    
    available_to_budget = max(cash_on_hand - all_budget, 0)
    actual_available_to_budget = max(actual_cash_on_hand - all_budget, 0)


    left_to_spend = balance + past_balance
    actual_left_to_spend = actual_balance + past_balance

    # actual_available_to_budget = max(actual_left_to_spend - all_budget - total_saved_so_far, 0)


    # Aggregating all expenses (full list, includes savings) by category
    category_expenses = defaultdict(float)
    for exp in expenses:
        if exp.subcategory and exp.subcategory.category:
            key = exp.subcategory.category.name
        elif exp.category:
            key = exp.category.name
        else:
            key = 'Uncategorized'
        category_expenses[key] += float(exp.amount)

    # Budget vs Actual table population
    for item in budget_items:
        key = item.subcategory.name if item.subcategory else item.category.name
        summary[key]['budgeted'] += float(item.forecasted_amount)

    for exp in expenses_for_budget_vs_actual:
        key = exp.subcategory.name if exp.subcategory else exp.category.name
        summary[key]['actual'] += float(exp.amount)

    #report = []
   # for key, data in summary.items():
     #   data['difference'] = data['budgeted'] - data['actual']
     #   report.append({
      #      'item': key,
      #      **data
      #  })

    report = []
    for key, data in summary.items():
        data['difference'] = data['budgeted'] - data['actual']
        
        # Only include if at least one > 0
        if data['budgeted'] > 0 or data['actual'] > 0:
            report.append({
                'item': key,
                **data
            })


    # Wallet data
    wallets = ExpenseSource.objects.filter(Q(user=user) | Q(user__isnull=True))
    wallet_data = []
    most_active_wallet = None

    for wallet in wallets:
        income_total = Income.objects.filter(
            user=user, income_source=wallet, date__range=[start_date, end_date]
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        expense_total = Expense.objects.filter(
            user=user, source=wallet, date__range=[start_date, end_date]
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        if income_total == 0 and expense_total == 0:
            continue

        balance_wallet = income_total - expense_total
        wallet_data.append({
            'name': wallet.name,
            'income_total': income_total,
            'expense_total': expense_total,
            'balance': balance_wallet,
        })

    if wallet_data:
        most_active_wallet = max(wallet_data, key=lambda w: float(w.get('income_total', 0)) + float(w.get('expense_total', 0)))

    # Subcategory aggregation for donut chart
    subcategory_expenses = defaultdict(float)
    for exp in expenses:
        if exp.subcategory:
            key = exp.subcategory.name
        else:
            key = 'Uncategorized'
        subcategory_expenses[key] += float(exp.amount)

    goals = SavingsGoal.objects.filter(user=user).order_by('-created_at')
    wallets = ExpenseSource.objects.filter(Q(user=user) | Q(user__is_superuser=True) | Q(user__isnull=True)).order_by('name')

    context = {
        'incomes': incomes,
        'expenses': expenses,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'report': report,
        'month': f"{start_date.strftime('%B')} â€“ {end_date.strftime('%B %Y')}" if start_date.month != end_date.month else start_date.strftime('%B %Y'),
        'total_budget': total_budget,
        'total_spent': total_spent,
        'budget_pct': budget_pct,
        'spent_pct': spent_pct,
        'category_expenses': category_expenses,
        'category_expenses_json': json.dumps(category_expenses),
        'subcategory_expenses': subcategory_expenses,
        'subcategory_expenses_json': json.dumps(subcategory_expenses),
        'income_minus_budget': income_minus_budget,
        'unassigned_budget': unassigned_budget,
        'wallet_data': wallet_data,
        'most_active_wallet': most_active_wallet,
        'recent_goals': SavingsGoal.objects.filter(user=user).order_by('-created_at')[:2],
        'past_balance': past_balance,
        'available_to_budget': available_to_budget,
        'left_to_spend': left_to_spend,
        'actual_left_to_spend': actual_left_to_spend,
        'actual_available_to_budget': actual_available_to_budget,
        'goals': goals,
        'wallets': wallets,
        'future_budget': future_budget,
        'planned_income_total': planned_income_total,
        'planned_expense_total': planned_expense_total,
    }

    return render(request, 'budget/my_dashboard.html', context)


@login_required
def top_up_savings(request):
    if request.method == 'POST':
        goal_id = request.POST.get('goal_id')
        amount_str = request.POST.get('amount')
        source_id = request.POST.get('source_id')

        goal = get_object_or_404(SavingsGoal, id=goal_id, user=request.user)
        try:
            amount = Decimal(amount_str)
            if amount <= 0:
                messages.warning(request, "Amount must be greater than zero.")
                return redirect('my_dashboard')

            # Optionally, check amount <= available_to_budget for this user

            goal.saved_so_far += amount
            goal.save()

            category, created = ExpenseCategory.objects.get_or_create(
                name="Savings",
                user=request.user,
                defaults={'is_savings_generated': True}
            )
            if not created and not category.is_savings_generated:
                category.is_savings_generated = True
                category.save()

            subcategory, _ = ExpenseSubCategory.objects.get_or_create(
                name=goal.subcategory.name,
                category=category,
                user=request.user
            )

            wallet = ExpenseSource.objects.filter(
                Q(id=source_id) & (Q(user=request.user) | Q(user__is_superuser=True) | Q(user__isnull=True))
            ).first()

            Expense.objects.create(
                user=request.user,
                category=category,
                subcategory=subcategory,
                description=f"Added to savings jar: {goal.subcategory}",
                amount=amount,
                source=wallet,
                date=timezone.now().date()
            )

            messages.success(request, f"KES {amount} successfully added to {goal.subcategory}.")
            return redirect('my_dashboard')

        except (InvalidOperation, TypeError):
            messages.error(request, "Invalid amount. Please enter a valid number.")
            return redirect('my_dashboard')

    return redirect('my_dashboard')


from django.core.paginator import Paginator

@login_required
def dashboard(request):
    user = request.user
    today = date.today()
    incomes_qs = Income.objects.filter(
        user=user,
        amount__gt=0
    ).order_by('-is_favourite', '-date')
    
    expenses_qs = Expense.objects.filter(
        user=user,
        amount__gt=0
    ).order_by('-is_favourite', '-date')


    for inc in incomes_qs:
        inc.is_upcoming = inc.date > today

    for exp in expenses_qs:
        exp.is_planned = exp.date > today

    # ✅ Paginator for incomes
    income_paginator = Paginator(incomes_qs, 20)  # 20 items per page
    income_page_number = request.GET.get("income_page")
    incomes = income_paginator.get_page(income_page_number)

    # ✅ Paginator for expenses
    expense_paginator = Paginator(expenses_qs, 20)  # 20 items per page
    expense_page_number = request.GET.get("expense_page")
    expenses = expense_paginator.get_page(expense_page_number)

    goals = SavingsGoal.objects.filter(user=user)

    planned_incomes_qs = Income.objects.filter(user=user, date__gt=today)
    planned_income_total = planned_incomes_qs.aggregate(total=Sum('amount'))['total'] or 0

    planned_expenses_qs = Expense.objects.filter(user=user, date__gt=today)
    planned_expense_total = planned_expenses_qs.aggregate(total=Sum('amount'))['total'] or 0

    # ---- Totals ----
    total_income = sum(i.amount for i in incomes_qs)
    total_expense = sum(e.amount for e in expenses_qs)
    balance = total_income - total_expense

    # ---- CATEGORY AGGREGATION ----
    category_expenses = defaultdict(float)
    for exp in expenses_qs:   # use full queryset, not paginated
        if exp.subcategory and exp.subcategory.category:
            key = exp.subcategory.category.name
        elif exp.category:
            key = exp.category.name
        else:
            key = 'Uncategorized'
        category_expenses[key] += float(exp.amount)

    # ---- SUBCATEGORY AGGREGATION ----
    subcategory_expenses = defaultdict(float)
    for exp in expenses_qs:   # use full queryset, not paginated
        if exp.subcategory:
            key = exp.subcategory.name
        else:
            key = 'Uncategorized'
        subcategory_expenses[key] += float(exp.amount)

    # Dates for averages
    dates = expenses_qs.aggregate(first_date=Min('date'), last_date=Max('date'))
    first_date, last_date = dates['first_date'], dates['last_date']

    if first_date and last_date:
        days = max((last_date - first_date).days, 1)
        months = max(Decimal(days) / Decimal(30), Decimal(1))
        average_daily_spending = total_expense / Decimal(days)
        average_monthly_spending = total_expense / months

        subcats = [exp.subcategory.name for exp in expenses_qs if exp.subcategory]
        most_common_subcat = Counter(subcats).most_common(1)
        most_frequent_subcategory = most_common_subcat[0][0] if most_common_subcat else 'N/A'
    else:
        average_daily_spending = 0
        average_monthly_spending = 0
        most_frequent_subcategory = 'N/A'

    # Largest subcategory outflow
    if subcategory_expenses:
        largest_subcategory, largest_outflow = max(subcategory_expenses.items(), key=itemgetter(1))
    else:
        largest_subcategory, largest_outflow = 'N/A', 0

    # Most frequent subcategory txn count
    most_freq_subcat_txn_count = 0
    if most_frequent_subcategory != 'N/A':
        most_freq_subcat_txn_count = expenses_qs.filter(subcategory__name=most_frequent_subcategory).count()
    
    # Income chart prep
    category_subcategory_data = defaultdict(lambda: defaultdict(float))
    for inc in incomes_qs:   # use full queryset
        category_name = inc.category.name if inc.category else "Uncategorized"
        subcategory_name = inc.subcategory.name if inc.subcategory else "Uncategorized"
        category_subcategory_data[category_name][subcategory_name] += float(inc.amount)

    categories = list(category_subcategory_data.keys())
    subcategories = sorted({sub for subs in category_subcategory_data.values() for sub in subs})

    datasets = []
    colors = [
        'rgba(54, 162, 235, 0.7)',
        'rgba(255, 99, 132, 0.7)',
        'rgba(255, 206, 86, 0.7)',
        'rgba(75, 192, 192, 0.7)',
        'rgba(153, 102, 255, 0.7)',
        'rgba(255, 159, 64, 0.7)',
        'rgba(199, 199, 199, 0.7)'
    ]
    for i, sub in enumerate(subcategories):
        datasets.append({
            'label': sub,
            'data': [category_subcategory_data[cat].get(sub, 0) for cat in categories],
            'backgroundColor': colors[i % len(colors)],
            'borderWidth': 1
        })

    context = {
        'incomes': incomes,  # now paginated
        'expenses': expenses,  # now paginated
        'goals': goals,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'category_expenses': category_expenses,
        'category_expenses_json': json.dumps(dict(category_expenses)),
        'subcategory_expenses': subcategory_expenses,
        'subcategory_expenses_json': json.dumps(dict(subcategory_expenses)),
        'average_daily_spending': average_daily_spending,
        'average_monthly_spending': average_monthly_spending,
        'most_frequent_subcategory': most_frequent_subcategory,
        'largest_subcategory': largest_subcategory,
        'largest_outflow': largest_outflow,
        'most_freq_subcat_txn_count': most_freq_subcat_txn_count,
        'planned_income_total': planned_income_total,
        'planned_expense_total': planned_expense_total,
        'stacked_income_labels': json.dumps(categories),
        'stacked_income_datasets': json.dumps(datasets),
    }
    return render(request, 'budget/dashboard.html', context)


# @login_required
# def dashboard(request):
#     user = request.user
#     today = date.today()
#     incomes = Income.objects.filter(user=user).order_by('-is_favourite', '-date')
#     expenses = Expense.objects.filter(user=user).order_by('-is_favourite', '-date')
#     goals = SavingsGoal.objects.filter(user=user)

#     planned_incomes_qs = Income.objects.filter(user=user, date__gt=today)
#     planned_income_total = planned_incomes_qs.aggregate(total=Sum('amount'))['total'] or 0

#     planned_expenses_qs = Expense.objects.filter(user=user, date__gt=today)
#     planned_expense_total = planned_expenses_qs.aggregate(total=Sum('amount'))['total'] or 0

#     total_income = sum(i.amount for i in incomes)
#     total_expense = sum(e.amount for e in expenses)
#     balance = total_income - total_expense

#     # ---- CATEGORY AGGREGATION ----
#     category_expenses = defaultdict(float)
#     for exp in expenses:
#         if exp.subcategory and exp.subcategory.category:
#             key = exp.subcategory.category.name
#         elif exp.category:
#             key = exp.category.name
#         else:
#             key = 'Uncategorized'
#         category_expenses[key] += float(exp.amount)

#     # ---- SUBCATEGORY AGGREGATION ----
#     subcategory_expenses = defaultdict(float)
#     for exp in expenses:
#         if exp.subcategory:
#             key = exp.subcategory.name
#         else:
#             key = 'Uncategorized'
#         subcategory_expenses[key] += float(exp.amount)

#     # Aggregate dates for average spending calculations
#     dates = expenses.aggregate(
#         first_date=Min('date'),
#         last_date=Max('date')
#     )
#     first_date = dates['first_date']
#     last_date = dates['last_date']

#     if first_date and last_date:
#         days = max((last_date - first_date).days, 1)
#         months = max(Decimal(days) / Decimal(30), Decimal(1))
#         average_daily_spending = total_expense / Decimal(days)
#         average_monthly_spending = total_expense / months


#         subcats = [exp.subcategory.name for exp in expenses if exp.subcategory]
#         most_common_subcat = Counter(subcats).most_common(1)
#         most_frequent_subcategory = most_common_subcat[0][0] if most_common_subcat else 'N/A'
#     else:
#         average_daily_spending = 0
#         average_monthly_spending = 0
#         most_frequent_subcategory = 'N/A'

#     # Find category with largest outflow safely
#     if subcategory_expenses:
#         largest_subcategory, largest_outflow = max(subcategory_expenses.items(), key=itemgetter(1))
#     else:
#         largest_subcategory, largest_outflow = 'N/A', 0

#     # Count transactions for most frequent subcategory
#     most_freq_subcat_txn_count = 0
#     if most_frequent_subcategory != 'N/A':
#         most_freq_subcat_txn_count = expenses.filter(subcategory__name=most_frequent_subcategory).count()
    

#     # Data structure: {category: {subcategory: total_amount}}
#     category_subcategory_data = defaultdict(lambda: defaultdict(float))
    
#     for inc in incomes:
#         category_name = inc.category.name if inc.category else "Uncategorized"
#         subcategory_name = inc.subcategory.name if inc.subcategory else "Uncategorized"
    
#         category_subcategory_data[category_name][subcategory_name] += float(inc.amount)
    
#     # Prepare for Chart.js
#     categories = list(category_subcategory_data.keys())
#     subcategories = sorted({sub for subs in category_subcategory_data.values() for sub in subs})
    
#     datasets = []
#     colors = [
#         'rgba(54, 162, 235, 0.7)',
#         'rgba(255, 99, 132, 0.7)',
#         'rgba(255, 206, 86, 0.7)',
#         'rgba(75, 192, 192, 0.7)',
#         'rgba(153, 102, 255, 0.7)',
#         'rgba(255, 159, 64, 0.7)',
#         'rgba(199, 199, 199, 0.7)'
#     ]
    
#     for i, sub in enumerate(subcategories):
#         datasets.append({
#             'label': sub,
#             'data': [category_subcategory_data[cat].get(sub, 0) for cat in categories],
#             'backgroundColor': colors[i % len(colors)],
#             'borderWidth': 1
#         })


#     context = {
#         'incomes': incomes,
#         'expenses': expenses,
#         'goals': goals,
#         'total_income': total_income,
#         'total_expense': total_expense,
#         'balance': balance,
#         'category_expenses': category_expenses,
#         'category_expenses_json': json.dumps(dict(category_expenses)),
#         'subcategory_expenses': subcategory_expenses,
#         'subcategory_expenses_json': json.dumps(dict(subcategory_expenses)),
#         'average_daily_spending': average_daily_spending,
#         'average_monthly_spending': average_monthly_spending,
#         'most_frequent_subcategory': most_frequent_subcategory,
#         'largest_subcategory': largest_subcategory,
#         'largest_outflow': largest_outflow,
#         'most_freq_subcat_txn_count': most_freq_subcat_txn_count,
#         'planned_income_total': planned_income_total,
#         'planned_expense_total': planned_expense_total,

#         'stacked_income_labels': json.dumps(categories),
#         'stacked_income_datasets': json.dumps(datasets),

#     }
#     return render(request, 'budget/dashboard.html', context)


def spending_power(request):
    user = request.user
    today = date.today()

    all_incomes = Income.objects.filter(user=user).order_by('date')
    all_expenses = Expense.objects.filter(user=user).order_by('date')
    
    incomes = Income.objects.filter(user=user, date__lte=today)
    expenses = Expense.objects.filter(user=user, date__lte=today)
    

    total_income = sum((i.amount for i in incomes), Decimal('0'))
    total_expense = sum((e.amount for e in expenses), Decimal('0'))
    current_balance = total_income - total_expense

    # === Cash Flow Lag: longest gap between consecutive transactions ===
    all_dates = sorted([i.date for i in incomes] + [e.date for e in expenses])
    if all_dates:
        gaps = [(all_dates[i] - all_dates[i-1]).days for i in range(1, len(all_dates))]
        cash_flow_lag_days = max(gaps) if gaps else 0
    else:
        cash_flow_lag_days = 0

    # === Essentials & Total Averages ===
    essentials_expenses = expenses.filter(category__name__iexact='Essentials')
    total_essentials = sum((e.amount for e in essentials_expenses), Decimal('0'))

    # Use number of days with actual expenses for daily averages
    essentials_days = essentials_expenses.values('date').distinct().count() or 1
    daily_essentials_avg = total_essentials / Decimal(essentials_days)

    expense_days = expenses.values('date').distinct().count() or 1
    daily_total_avg = total_expense / Decimal(expense_days)

    days_cash_on_hand = (current_balance / daily_essentials_avg).quantize(Decimal('0.1')) if daily_essentials_avg > 0 else Decimal('0')
    days_cash_on_hand = max(days_cash_on_hand, Decimal('0'))

    liquidity_buffer_days = (current_balance / daily_total_avg).quantize(Decimal('0.1')) if daily_total_avg > 0 else Decimal('0')
    liquidity_buffer_days = max(liquidity_buffer_days, Decimal('0'))

    # === Monthly Trend Data ===
    monthly_data = defaultdict(lambda: {"income": Decimal('0'), "expense": Decimal('0'), "rollover": Decimal('0'), "date": None})
    for inc in incomes:
        m = inc.date.strftime("%Y-%m")
        monthly_data[m]["income"] += inc.amount
        if not monthly_data[m]["date"] or inc.date < monthly_data[m]["date"]:
            monthly_data[m]["date"] = inc.date
    for exp in expenses:
        m = exp.date.strftime("%Y-%m")
        monthly_data[m]["expense"] += exp.amount
        if not monthly_data[m]["date"] or exp.date < monthly_data[m]["date"]:
            monthly_data[m]["date"] = exp.date

    essentials_monthly_data = defaultdict(lambda: Decimal('0'))
    for e in essentials_expenses:
        m = e.date.strftime("%Y-%m")
        essentials_monthly_data[m] += e.amount

    sorted_months = sorted(monthly_data.keys())
    current_month_str = today.strftime("%Y-%m")

    # === Rollover ===
    prev_balance = Decimal('0')
    rollover_entries = []
    for m in sorted_months:
        month_rollover = prev_balance  # can be negative
        monthly_data[m]["rollover"] = max(month_rollover, Decimal('0'))  # for display/UI

        if month_rollover > 0 and monthly_data[m]["date"]:
            age_days = (today - monthly_data[m]["date"]).days
            rollover_entries.append((month_rollover, age_days))
            monthly_data[m]["age"] = age_days
        else:
            monthly_data[m]["age"] = 0

        prev_balance += monthly_data[m]["income"] - monthly_data[m]["expense"]

    # Weighted age of money
    positive_rollovers = [(amt, days) for amt, days in rollover_entries if amt > 0]
    total_rollover_amount = sum(amt for amt, _ in positive_rollovers)
    total_weighted_value = sum(amt * days for amt, days in positive_rollovers)
    weighted_age_today = round(float(total_weighted_value / total_rollover_amount), 1) if total_rollover_amount > 0 else 0

    # === Chart Data ===
    months_labels = [calendar.month_abbr[int(m.split("-")[1])] for m in sorted_months] or ["No Data"]
    liquidity_data, days_cash_data, age_of_money_data = [], [], []
    prev_balance = Decimal('0')

    for m in sorted_months:
        income = monthly_data[m]["income"]
        expense = monthly_data[m]["expense"]
        prev_balance += income - expense

        year, month_num = map(int, m.split("-"))
        days_for_avg = today.day if m == current_month_str else monthrange(year, month_num)[1]
        days_for_avg = max(days_for_avg, 1)

        avg_daily_total = (expense / Decimal(days_for_avg)) if expense > 0 else Decimal('0')
        liquidity_days = (prev_balance / avg_daily_total).quantize(Decimal('0.1')) if avg_daily_total > 0 else Decimal('0')
        liquidity_data.append(float(max(liquidity_days, Decimal('0'))))

        essentials_expense = essentials_monthly_data.get(m, Decimal('0'))
        avg_daily_essentials = (essentials_expense / Decimal(days_for_avg)) if essentials_expense > 0 else Decimal('0')
        days_cash = (prev_balance / avg_daily_essentials).quantize(Decimal('0.1')) if avg_daily_essentials > 0 else Decimal('0')
        days_cash_data.append(float(max(days_cash, Decimal('0'))))

        age_of_money_data.append(monthly_data[m]["age"])

    # Ensure no empty lists
    if not liquidity_data: liquidity_data = [0]
    if not days_cash_data: days_cash_data = [0]
    if not age_of_money_data: age_of_money_data = [0]

    # === Monthly Table ===
    monthly_table = []
    running_balance = Decimal('0')
    for i, m in enumerate(sorted_months):
        year, month_num = map(int, m.split("-"))
        days_in_month = monthrange(year, month_num)[1]
        expense_for_month = monthly_data[m]["expense"]
        income_for_month = monthly_data[m]["income"]

        running_balance += income_for_month - expense_for_month
        days_for_avg = today.day if m == current_month_str else days_in_month
        days_for_avg = max(days_for_avg, 1)

        avg_daily_spending = (expense_for_month / Decimal(days_for_avg)) if expense_for_month > 0 else Decimal('0')
        essentials_expense = essentials_monthly_data.get(m, Decimal('0'))
        avg_daily_essentials = (essentials_expense / Decimal(days_for_avg)) if essentials_expense > 0 else Decimal('0')

        liquidity_buffer = round(float(running_balance / avg_daily_spending), 2) if avg_daily_spending > 0 else 0
        liquidity_buffer = max(liquidity_buffer, 0)

        days_cash_on_hand_month = round(float(running_balance / avg_daily_essentials), 2) if avg_daily_essentials > 0 else 0
        days_cash_on_hand_month = max(days_cash_on_hand_month, 0)

        monthly_table.append((
            months_labels[i],
            round(float(monthly_data[m]["rollover"]), 2),
            round(float(avg_daily_spending), 2),
            monthly_data[m]["age"],
            liquidity_buffer,
            days_cash_on_hand_month,
        ))

    # === Current month rollover & percentage ===
    current_rollover_amount = monthly_data.get(current_month_str, {}).get("rollover", Decimal('0'))
    rollover_pct = round(float(current_rollover_amount / current_balance * 100), 1) if current_balance > 0 and current_rollover_amount > 0 else 0

    context = {
        'current_balance': float(current_balance),
        'cash_flow_lag_days': cash_flow_lag_days,
        'liquidity_buffer_days': float(liquidity_buffer_days),
        'days_cash_on_hand': float(days_cash_on_hand),
        'weighted_age_today': weighted_age_today,
        'months_labels': json.dumps(months_labels),
        'liquidity_data': json.dumps(liquidity_data),
        'days_cash_data': json.dumps(days_cash_data),
        'age_of_money_data': json.dumps([float(x) for x in age_of_money_data]),
        'monthly_table': monthly_table,
        'rollover_data': json.dumps([float(round(monthly_data[m]["rollover"], 2)) for m in sorted_months]),
        'new_income_data': json.dumps([float(round(max(monthly_data[m]["income"] - monthly_data[m]["rollover"], Decimal('0')), 2)) for m in sorted_months]),
        'rollover_amount': float(current_rollover_amount),
        'rollover_pct': rollover_pct,
    }

    return render(request, 'budget/spending_power.html', context)

import re
from django.db.models import Q

def normalize_name(name: str) -> str:
    # Lowercase
    name = name.lower()
    # Replace hyphens, underscores, slashes etc. with spaces
    name = re.sub(r'[-_/]+', ' ', name)
    # Collapse multiple spaces into one
    name = re.sub(r'\s+', ' ', name)
    # Strip leading/trailing spaces
    return name.strip()


############################################################################################ EXPENSES #########################################################################################################################################


@login_required
def expense_list(request):
    expenses = Expense.objects.filter(user=request.user).order_by('-date')

    return render(request, 'budget/expense_list.html', {
        'expenses': expenses
    })

from .utils import get_user_subcategories
from .models import UserHiddenSubCategory, UserHiddenIncomeSubCategory, UserHiddenWallet

@login_required 
def add_expense(request):
    user = request.user
    form = ExpenseForm(request.POST or None, user=user)

    if request.method == 'POST':
        if form.is_valid():
            expense = form.save(commit=False)
            expense.user = user
            expense.category = expense.subcategory.category  # 👈 auto-fill parent
            expense.save()

            wallet = expense.source  # This is the ExpenseSource

            total_income = wallet.income_set.filter(user=user).aggregate(total=models.Sum('amount'))['total'] or 0
            total_expense = wallet.expense_set.filter(user=user).aggregate(total=models.Sum('amount'))['total'] or 0
            balance = total_income - total_expense

            if balance < 0:
                messages.warning(request, f"Expense added, but the wallet '{wallet.name}' is now negative (KES {balance:.2f}).")
            else:
                messages.success(request, f"Expense of KES {expense.amount} for {expense.subcategory} added successfully.")

            return redirect('my_dashboard')
        else:
            messages.warning(request, "Please make sure all required fields are filled.")

            
        # Do not change sub...
    categories = ExpenseCategory.objects.filter(
        (Q(user__isnull=True) | Q(user__is_superuser=True))
    )
    #subcategories = ExpenseSubCategory.objects.filter(
    #    (Q(user__isnull=True) | Q(user=user) | Q(user__is_superuser=True))
    #)

    # Use helper that excludes hidden subcategories
    subcategories = get_user_subcategories(user)
    
    sources = ExpenseSource.objects.filter(
        Q(user__isnull=True) | Q(user=user) | Q(user__is_superuser=True)
    )

    expenses = Expense.objects.filter(user=user).order_by('-date')[:5]

    return render(request, 'budget/add_expense.html', {
        'form': form,
        'categories': categories,
        'subcategories': subcategories,
        'sources': sources,
        'expenses': expenses,
    })

@login_required
def manage_subcategories(request):
    # --- Fetch all subcategories (global + user) ---
    expense_subcategories = ExpenseSubCategory.objects.filter(
        Q(user=request.user) | Q(user__isnull=True)
    ).select_related('category').order_by('category__name', 'name')

    income_subcategories = IncomeSubCategory.objects.filter(
        Q(user=request.user) | Q(user__isnull=True)
    ).select_related('category').order_by('category__name', 'name')

    hidden_expense_ids = list(
        UserHiddenSubCategory.objects.filter(user=request.user)
        .values_list('subcategory_id', flat=True)
    )
    hidden_income_ids = list(
        UserHiddenIncomeSubCategory.objects.filter(user=request.user)
        .values_list('subcategory_id', flat=True)
    )

    # --- Handle Reset All Hidden ---
    if request.method == 'POST' and 'reset_hidden' in request.POST:
        UserHiddenSubCategory.objects.filter(user=request.user).delete()
        UserHiddenIncomeSubCategory.objects.filter(user=request.user).delete()
        messages.success(request, "All hidden subcategories have been reset.")
        return redirect('manage_subcategories')

    # --- Handle Save Visible (reverse logic) ---
    if request.method == 'POST' and 'save_hidden' in request.POST:
        hidden_type = request.POST.get('hidden_type')  # "income" or "expense"
        selected_ids = list(map(int, request.POST.getlist('visible_subcategories')))

        if hidden_type == 'expense':
            # All possible expense subcategories
            all_expense_ids = list(
                ExpenseSubCategory.objects.filter(
                    Q(user=request.user) | Q(user__isnull=True)
                ).values_list('id', flat=True)
            )

            # Anything NOT selected is hidden
            to_hide = set(all_expense_ids) - set(selected_ids)

            # Clear and recreate hidden entries
            UserHiddenSubCategory.objects.filter(user=request.user).delete()
            for sub_id in to_hide:
                UserHiddenSubCategory.objects.get_or_create(
                    user=request.user, subcategory_id=sub_id
                )

            messages.success(request, "Expenses updated successfully.")

        elif hidden_type == 'income':
            # All possible income subcategories
            all_income_ids = list(
                IncomeSubCategory.objects.filter(
                    Q(user=request.user) | Q(user__isnull=True)
                ).values_list('id', flat=True)
            )

            # Anything NOT selected is hidden
            to_hide = set(all_income_ids) - set(selected_ids)

            # Clear and recreate hidden entries
            UserHiddenIncomeSubCategory.objects.filter(user=request.user).delete()
            for sub_id in to_hide:
                UserHiddenIncomeSubCategory.objects.get_or_create(
                    user=request.user, subcategory_id=sub_id
                )

            messages.success(request, "Incomes updated successfully.")

        return redirect('manage_subcategories')

    expense_categories = ExpenseCategory.objects.filter(Q(user=request.user) | Q(user__isnull=True))
    income_categories = IncomeCategory.objects.filter(Q(user=request.user) | Q(user__isnull=True))

    # --- Render Template ---
    context = {
        'expense_subcategories': expense_subcategories,
        'income_subcategories': income_subcategories,
        'hidden_expense_ids': hidden_expense_ids,
        'hidden_income_ids': hidden_income_ids,
        'expense_categories': expense_categories,
        'income_categories': income_categories,        
    }

    return render(request, 'budget/manage_subcategories.html', context)


@login_required
def manage_wallets(request):
    user = request.user

    # All wallets (system + user)
    wallets = ExpenseSource.objects.filter(
        Q(user=user) | Q(user__isnull=True)
    ).order_by('name')

    all_wallet_ids = list(wallets.values_list('id', flat=True))

    # Current hidden wallets for this user
    hidden_wallet_ids = list(
        UserHiddenWallet.objects.filter(user=user, hidden=True)
        .values_list('wallet_id', flat=True)
    )

    # Reset hidden wallets
    if request.method == 'POST' and 'reset_hidden' in request.POST:
        UserHiddenWallet.objects.filter(user=user, wallet_id__in=all_wallet_ids).update(hidden=False)
        messages.success(request, "All hidden wallets have been reset.")
        return redirect('manage_wallets')

    # Save changes (checked = visible)
    if request.method == 'POST' and 'save_hidden' in request.POST:
        visible_ids_raw = request.POST.getlist('visible_wallets')
        visible_ids = set(map(int, visible_ids_raw))
        to_hide = set(all_wallet_ids) - visible_ids

        for wid in all_wallet_ids:
            UserHiddenWallet.objects.update_or_create(
                user=user,
                wallet_id=wid,
                defaults={'hidden': wid in to_hide}
            )

        messages.success(request, "Wallets updated successfully.")
        return redirect('manage_wallets')

    context = {
        'wallets': wallets,
        'hidden_wallet_ids': hidden_wallet_ids,
    }
    return render(request, 'budget/manage_wallets.html', context)



@login_required
def edit_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)

    if request.method == 'POST':
        form = ExpenseForm(request.POST, instance=expense, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Expense edited successfully.")
            return redirect('dashboard')
        else:
            messages.warning(request, "Please make sure all required fields are filled.")

    else:
        form = ExpenseForm(instance=expense, user=request.user)

    categories = ExpenseCategory.objects.filter(
        models.Q(user=request.user) | models.Q(user__isnull=True)
    )
    #subcategories = ExpenseSubCategory.objects.filter(
    #    models.Q(user=request.user) | models.Q(user__isnull=True)
    #)

    # Use helper that excludes hidden subcategories
    subcategories = get_user_subcategories(request.user)

    return render(request, 'budget/add_expense.html', {
        'form': form,
        'editing': True,
        'expense': expense,
        'categories': categories,
        'subcategories': subcategories,
    })


@login_required
def delete_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    
    if request.method == 'POST':
        expense.delete()
        messages.success(request, "Expense deleted successfully.")
        return redirect('dashboard')

    return render(request, 'budget/delete_expense.html', {'expense': expense})


from django.http import JsonResponse

import json

@login_required
def add_expense_source(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            raw_name = data.get('name', '')
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})

        if raw_name.strip():
            normalized = normalize_name(raw_name)

            # Search in both user's and admin/global sources
            existing = ExpenseSource.objects.filter(
                models.Q(user=request.user) | models.Q(user__isnull=True)
            )
            for source in existing:
                if normalize_name(source.name) == normalized:
                    return JsonResponse({'success': False, 'error': 'This wallet already exists'})

            # Save a "pretty" version
            source = ExpenseSource.objects.create(
                user=request.user,
                name=raw_name.strip().title()
            )

            return JsonResponse({'success': True, 'id': source.id, 'name': source.name})

    return JsonResponse({'success': False, 'error': 'Invalid or missing source name'})

@login_required
def add_category(request):
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.user = request.user
            category.save()
            return redirect('add_expense')  # after saving, go back to the expense page
    else:
        form = ExpenseCategoryForm()

    return render(request, 'budget/add_category.html', {'form': form})  # not used here but still needed


@csrf_exempt
@login_required
def add_subcategory(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})

        name = data.get('name', '')
        category_id = data.get('category_id')

        if not name.strip() or not category_id:
            return JsonResponse({'success': False, 'error': 'Missing name or category'})

        category = get_object_or_404(ExpenseCategory, id=category_id)
        normalized = normalize_name(name)

        # Check both user-created and admin/global subcategories
        existing = ExpenseSubCategory.objects.filter(
            category=category
        ).filter(
            models.Q(user=request.user) | models.Q(user__isnull=True)
        )

        for sub in existing:
            if normalize_name(sub.name) == normalized:
                return JsonResponse({'success': False, 'error': 'This subcategory already exists.'})

        subcategory = ExpenseSubCategory.objects.create(
            name=name.strip().title(),
            category=category,
            user=request.user
        )

        return JsonResponse({'success': True, 'id': subcategory.id, 'name': subcategory.name})

    return JsonResponse({'success': False, 'error': 'Invalid request'})



############################################################################################ INCOMES #########################################################################################################################################

@login_required
def income_list(request):
    incomes = Income.objects.filter(user=request.user).order_by('-date')

    return render(request, 'budget/income_list.html', {
        'incomes': incomes
    })


@login_required
def add_income(request):
    user = request.user
    if request.method == 'POST':
        form = IncomeForm(request.POST, user=user)
        if form.is_valid():
            income = form.save(commit=False)
            income.user = user
            income.category = income.subcategory.category  # 👈 auto-fill parent category
            income.save()

            wallet = income.income_source
            total_income = wallet.income_set.filter(user=user).aggregate(total=models.Sum('amount'))['total'] or 0
            total_expense = wallet.expense_set.filter(user=user).aggregate(total=models.Sum('amount'))['total'] or 0
            balance = total_income - total_expense

            if balance < 0:
                messages.warning(request, f"Income added, but the wallet '{wallet.name}' is still negative (KES {balance:.2f}).")
            else:
                messages.success(request, f"Income of KES {income.amount} from {income.subcategory} added successfully.")

            return redirect('my_dashboard')
        else:
            messages.warning(request, "Please make sure all required fields are filled.")
    else:
        form = IncomeForm(user=user)

    # --- Hidden income subcategories ---
    hidden_income_ids = UserHiddenIncomeSubCategory.objects.filter(
        user=user, hidden=True
    ).values_list('subcategory_id', flat=True)


    # these only matter if you want dropdowns elsewhere
    excluded_categories = ["Investment Income"]
    categories = IncomeCategory.objects.filter(
        models.Q(user=user) | models.Q(user__isnull=True)
    ).exclude(name__in=excluded_categories)

    subcategories = IncomeSubCategory.objects.filter(
        models.Q(user=user) | models.Q(user__isnull=True)
    ).exclude(id__in=hidden_income_ids)

    return render(request, 'budget/add_income.html', {
        'form': form,
        'categories': categories,  # optional now
        'subcategories': subcategories
    })


@login_required
def edit_income(request, income_id):
    user = request.user
    income = get_object_or_404(Income, id=income_id, user=user)

    if request.method == 'POST':
        form = IncomeForm(request.POST, instance=income, user=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Income edited successfully.")
            return redirect('my_dashboard')
        else:
            messages.warning(request, "Please make sure all required fields are filled.")
    else:
        form = IncomeForm(instance=income, user=user)

    # --- Hidden income subcategories ---
    hidden_income_ids = UserHiddenIncomeSubCategory.objects.filter(
        user=user, hidden=True
    ).values_list('subcategory_id', flat=True)

    # --- Hidden wallets ---
    hidden_wallet_ids = UserHiddenWallet.objects.filter(
        user=user, hidden=True
    ).values_list('wallet_id', flat=True)

    # --- Wallets (excluding hidden + system wallets) ---
    wallets = ExpenseSource.objects.filter(
        models.Q(user__isnull=True) | models.Q(user=user) | models.Q(user__is_superuser=True)
    ).exclude(
        id__in=hidden_wallet_ids
    )

    # --- Excluded income categories ---
    excluded_categories = ["Investment Income"]
    categories = IncomeCategory.objects.filter(
        models.Q(user=user) | models.Q(user__isnull=True)
    ).exclude(name__in=excluded_categories)

    # --- Income subcategories (excluding hidden) ---
    subcategories = IncomeSubCategory.objects.filter(
        models.Q(user=user) | models.Q(user__isnull=True)
    ).exclude(id__in=hidden_income_ids)

    return render(request, 'budget/add_income.html', {
        'form': form,
        'editing': True,
        'income': income,
        'categories': categories,
        'subcategories': subcategories,
        'wallets': wallets,
    })



@login_required
def delete_income(request, income_id):
    income = get_object_or_404(Income, id=income_id, user=request.user)
    if request.method == 'POST':
        income.delete()
        messages.success(request, "Income deleted successfully.")
        return redirect('dashboard')


from django.db.models import Q

@login_required
def add_income_subcategory(request):
    if request.method == 'POST':
        category_id = request.POST.get('category')
        name = request.POST.get('name')

        if category_id and name:
            category = get_object_or_404(IncomeCategory, id=category_id)

            # Normalize for consistent duplicate detection
            normalized = normalize_name(name)

            # Check both user-created AND admin/global subcategories
            existing = IncomeSubCategory.objects.filter(
                category=category
            ).filter(
                Q(user=request.user) | Q(user__isnull=True)  # include admin/global
            )

            for sub in existing:
                if normalize_name(sub.name) == normalized:
                    return JsonResponse({'error': 'This subcategory already exists.'}, status=400)

            # Store the "pretty" version (title case)
            subcat = IncomeSubCategory.objects.create(
                category=category,
                name=name.strip().title(),
                user=request.user
            )

            return JsonResponse({
                'id': subcat.id,
                'name': subcat.name,
                'category_id': category.id,
            })

    return JsonResponse({'error': 'Invalid data'}, status=400)
   
############################################################################################ BUDGETS #########################################################################################################################################


@login_required
def all_budgets_view(request):
    budgets = BudgetItem.objects.filter(user=request.user).order_by('-month')

    return render(request, 'budget/all_budgets.html', {
        'budgets': budgets,
    })



@login_required
def all_budget_vs_actual_view(request):
    user = request.user

    budgets = BudgetItem.objects.filter(user=user).order_by('-is_favourite', '-month')

    # Get all months with budgets or expenses
    months = set(BudgetItem.objects.filter(user=user).values_list('month', flat=True))
    months.update(Expense.objects.filter(user=user).dates('date', 'month'))

    # Normalize to first day of each month
    months = sorted(set(m.replace(day=1) for m in months), reverse=True)

    monthly_reports = []

    for month in months:
        start = month
        end = month.replace(day=monthrange(month.year, month.month)[1])

        budget_items = BudgetItem.objects.filter(user=user, month__range=[start, end])

        #expenses = Expense.objects.filter(user=user, date__range=[start, end])

        expenses = Expense.objects.filter(
            user=user, date__range=[start, end]
        ).exclude(category__isnull=True, subcategory__isnull=True)


        summary = defaultdict(lambda: {'budgeted': 0, 'actual': 0})

        total_budget = 0
        total_spent = 0

        # Aggregate budgeted amounts
        for item in budget_items:
            key = item.subcategory.name if item.subcategory else item.category.name
            summary[key]['budgeted'] += float(item.forecasted_amount)
            total_budget += float(item.forecasted_amount)

        # Aggregate actual spending
        for exp in expenses:
            key = exp.subcategory.name if exp.subcategory else exp.category.name
            summary[key]['actual'] += float(exp.amount)
            total_spent += float(exp.amount)

        remaining_budget = max(total_budget - total_spent, 0)

        # Build table report
        #report = []
        #for key, data in summary.items():
        #    data['difference'] = data['budgeted'] - data['actual']
        #    report.append({
        #        'item': key,
        #        **data
        #    })

        report = []
        for key, data in summary.items():
            data['difference'] = data['budgeted'] - data['actual']
            
            # Only include if at least one > 0
            if data['budgeted'] > 0 or data['actual'] > 0:
                report.append({
                    'item': key,
                    **data
                })


        monthly_reports.append({
            'month': month.strftime('%B %Y'),
            'report': report,
            'total_budget': total_budget,
            'total_spent': total_spent,
            'remaining_budget': remaining_budget,
        })

    return render(request, 'budget/all_budget_vs_actual.html', {
        'budgets': budgets,
        'monthly_reports': monthly_reports
    })



@login_required
def set_budget(request, pk=None):
    if pk:
        budget_item = get_object_or_404(BudgetItem, pk=pk, user=request.user)
        editing = True
    else:
        budget_item = None
        editing = False

    duplicate_warning = None

    if request.method == 'POST':
        form = BudgetItemForm(request.POST, instance=budget_item, user=request.user)
        if form.is_valid():
            new_item = form.save(commit=False)
            new_item.user = request.user

            duplicate = BudgetItem.objects.filter(
                user=request.user,
                month=new_item.month,
                subcategory__name__iexact=new_item.subcategory.name
            ).exclude(pk=new_item.pk).exists()

            if duplicate:
                messages.warning(request, f"A budget for '{new_item.subcategory}' already exists for {new_item.month.strftime('%B %Y')}.")
                duplicate_warning = f"A budget for '{new_item.subcategory}' already exists for {new_item.month.strftime('%B %Y')}."
                return redirect('all_budget_vs_actual')
            else:
                new_item.save()
                messages.success(request, 'Budget item updated successfully.' if editing else 'Budget item saved successfully.')
                return redirect('all_budget_vs_actual')
        else:
            messages.warning(request, "Please make sure all required fields are filled correctly.")
    else:
        form = BudgetItemForm(instance=budget_item, user=request.user)

    current_month = date.today().replace(day=1)
    budgets = BudgetItem.objects.filter(user=request.user, month=current_month)

    categories = ExpenseCategory.objects.filter(
        (
            Q(user__isnull=True) | Q(user__is_superuser=True)
        ) &
        (
            Q(is_savings_generated=False) | Q(is_savings_generated__isnull=True)
        )
    )

    return render(request, 'budget/set_budget.html', {
        'form': form,
        'budgets': budgets,
        'editing': editing,
        'categories': categories,
        'duplicate_warning': duplicate_warning
    })




@login_required
def edit_budget(request, pk):
    budget_item = get_object_or_404(BudgetItem, pk=pk, user=request.user)

    if request.method == 'POST':
        form = BudgetItemForm(request.POST, instance=budget_item)
        if form.is_valid():
            form.save()
            messages.success(request, 'Budget item updated.')
            return redirect('all_budget_vs_actual')
    else:
        form = BudgetItemForm(instance=budget_item)

    return render(request, 'budget/edit_budget.html', {'form': form})


@login_required
def delete_budget(request, pk):
    item = get_object_or_404(BudgetItem, pk=pk, user=request.user)
    if request.method == 'POST':
        item.delete()
        messages.success(request, 'Budget deleted successfully.')
    return redirect('all_budget_vs_actual')

@login_required
def budget_vs_actual_view(request):
    user = request.user
    today = date.today()
    current_month_start = today.replace(day=1)
    current_month_end = today.replace(day=28) + timedelta(days=4)
    current_month_end = current_month_end.replace(day=1) - timedelta(days=1)

    # Fetching budget items and expenses for current month
    budget_items = BudgetItem.objects.filter(
        user=user, month__range=[current_month_start, current_month_end]
    )
    expenses = Expense.objects.filter(
        user=user, date__range=[current_month_start, current_month_end]
    )

    summary = defaultdict(lambda: {'budgeted': 0, 'actual': 0})

    # Filling in budgeted amounts
    for item in budget_items:
        key = item.subcategory.name if item.subcategory else item.category.name
        summary[key]['budgeted'] += float(item.forecasted_amount)

    # Filling in actual expenses
    for exp in expenses:
        key = exp.subcategory.name if exp.subcategory else exp.category.name
        summary[key]['actual'] += float(exp.amount)

    # Add difference
    report = []
    for key, data in summary.items():
        data['difference'] = data['budgeted'] - data['actual']
        report.append({
            'item': key,
            **data,
        })

    context = {
        'report': report,
        'month': today.strftime('%B %Y'),
    }

    return render(request, 'budget/budget_vs_actual.html', context)


############################################################################################ SAVINGS #########################################################################################################################################


@login_required
def saving_list(request):
    goals = SavingsGoal.objects.filter(user=request.user).order_by('-created_at')

    # Get savings-generated transactions for each goal
    savings_data = []
    for goal in goals:
        transactions = Expense.objects.filter(
            user=request.user,
            subcategory=goal.subcategory,
            category__is_savings_generated=True,
            amount__gt=0
        ).order_by('-date')

        savings_data.append({
            'goal': goal,
            'transactions': transactions
        })

    return render(request, 'budget/saving_list.html', {
        'goals': goals,
        'savings_data': savings_data
    })


@login_required
def create_goal(request):
    # Ensure the "Savings" category exists
    ExpenseCategory.objects.get_or_create(
        name="Savings",
        user=request.user,
        defaults={"is_savings_generated": True}
    )

    if request.method == 'POST':
        form = SavingsGoalForm(request.POST, user=request.user)
        if form.is_valid():
            goal = form.save(commit=False)
            goal.user = request.user

            # Prevent duplicate jar names
            duplicate = SavingsGoal.objects.filter(
                user=request.user,
                subcategory__name__iexact=goal.subcategory.name
            ).exclude(pk=goal.pk).exists()
            if duplicate:
                messages.warning(request, f"A Savings Jar for '{goal.subcategory}' already exists. Please see below.")
                return redirect('saving_list')

            # Ensure its category is marked as savings-generated
            category = goal.subcategory.category
            if not category.is_savings_generated:
                category.is_savings_generated = True
                category.save()

            goal.save()

            # Pick a wallet if provided
            source_id = request.POST.get('source')
            wallet = ExpenseSource.objects.filter(
                Q(id=source_id) & (Q(user=request.user) | Q(user__is_superuser=True) | Q(user__isnull=True))
            ).first()

            # 🔹 Always create a zero-expense entry (so it shows up in budgets)
            Expense.objects.create(
                user=request.user,
                category=goal.subcategory.category,
                subcategory=goal.subcategory,
                description=f"Savings Jar created: {goal.subcategory.name}",
                amount=0,
                source=wallet,
                date=timezone.now().date()
            )

            # 🔹 If initial savings entered, record it as another expense
            if goal.saved_so_far > 0:
                Expense.objects.create(
                    user=request.user,
                    category=goal.subcategory.category,
                    subcategory=goal.subcategory,
                    description=f"Initial savings for: {goal.subcategory.name}",
                    amount=goal.saved_so_far,
                    source=wallet,
                    date=goal.deadline or timezone.now().date()
                )

            messages.success(request, f"Savings Jar for {goal.subcategory} created successfully.")
            return redirect('saving_list')

        else:
            messages.warning(request, "Please make sure all required fields are filled correctly.")
    else:
        form = SavingsGoalForm(user=request.user)

    return render(request, 'budget/create_goal.html', {'form': form})



@login_required
def add_to_goal(request, goal_id):
    goal = SavingsGoal.objects.get(id=goal_id, user=request.user)

    if request.method == 'POST':
        amount = Decimal(request.POST.get('amount'))
        goal.saved_so_far += amount
        goal.save()
        return redirect('dashboard')

    return render(request, 'budget/add_to_goal.html', {'goal': goal})



@login_required
def edit_goal(request, goal_id):
    goal = get_object_or_404(SavingsGoal, id=goal_id, user=request.user)
    user = request.user

    if request.method == 'POST':
        if request.POST.get("add_to_goal"):
            # ✅ Handle top-up
            try:
                amount = Decimal(request.POST.get('amount'))
                if amount > 0:
                    goal.saved_so_far += amount
                    goal.save()

                    # ✅ Ensure Savings category exists
                    category, created = ExpenseCategory.objects.get_or_create(
                        name="Savings",
                        user=user,
                        defaults={'is_savings_generated': True}
                    )
                    if not created and not category.is_savings_generated:
                        category.is_savings_generated = True
                        category.save()

                    # ✅ Match subcategory to the goal’s subcategory
                    subcategory, _ = ExpenseSubCategory.objects.get_or_create(
                        name=goal.subcategory.name,
                        category=category,
                        user=user
                    )

                    # ✅ Get wallet (exclude hidden ones)
                    source_id = request.POST.get('source_id')

                    hidden_wallet_ids = UserHiddenWallet.objects.filter(
                        user=user, hidden=True
                    ).values_list('wallet_id', flat=True)

                    wallet = ExpenseSource.objects.filter(
                        Q(id=source_id),
                        Q(user=user) | Q(user__is_superuser=True) | Q(user__isnull=True)
                    ).exclude(
                        id__in=hidden_wallet_ids
                    ).first()

                    # ✅ Record as expense
                    Expense.objects.create(
                        user=user,
                        category=category,
                        subcategory=subcategory,
                        description=f"Added to savings jar: {goal.subcategory}",
                        amount=amount,
                        source=wallet,
                        date=timezone.now().date()
                    )

                    messages.success(request, f"KES {amount} successfully added to {goal.subcategory}.")
                else:
                    messages.warning(request, "Amount must be greater than zero.")
                return redirect('edit_goal', goal_id=goal.id)

            except (InvalidOperation, TypeError):
                messages.error(request, "Invalid amount. Please enter a valid number.")

        else:
            # ✅ Handle goal edit form
            form = SavingsGoalForm(request.POST, instance=goal, user=user)
            if form.is_valid():
                form.save()
                messages.success(request, "Goal updated successfully.")
                return redirect('saving_list')
            else:
                messages.error(request, "Please correct the errors below.")

    else:
        form = SavingsGoalForm(instance=goal, user=user)

    # ✅ Hidden wallets
    hidden_wallet_ids = UserHiddenWallet.objects.filter(
        user=user, hidden=True
    ).values_list('wallet_id', flat=True)

    # ✅ Wallets (excluding hidden ones)
    wallets = ExpenseSource.objects.filter(
        Q(user=user) | Q(user__is_superuser=True) | Q(user__isnull=True)
    ).exclude(
        id__in=hidden_wallet_ids
    ).order_by('name')

    return render(request, 'budget/create_goal.html', {
        'form': form,
        'editing': True,
        'goal': goal,
        'wallets': wallets,
    })


@login_required
def delete_goal(request, goal_id):
    goal = get_object_or_404(SavingsGoal, id=goal_id, user=request.user)

    if request.method == 'POST':
        goal.delete()
        messages.success(request, 'Savings withdrawn successfully.')

        return redirect('saving_list')

    return render(request, 'budget/confirm_delete_goal.html', {'goal': goal})

@login_required
@csrf_exempt
def add_savings_subcategory(request):
    if request.method == "POST":
        name = request.POST.get("name")
        if name:
            category, _ = ExpenseCategory.objects.get_or_create(
                name="Savings",
                user=request.user
            )
            subcat, created = ExpenseSubCategory.objects.get_or_create(
                name=name,
                category=category,
                user=request.user
            )
            return JsonResponse({"success": True, "id": subcat.id, "name": subcat.name})
        else:
            return JsonResponse({"success": False, "error": "Name required."})
    return JsonResponse({"success": False, "error": "Invalid request method."})


############################################################################################ REPORTS #########################################################################################################################################
@login_required
def custom_report(request):
    user = request.user
    today = date.today()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        try:
            # use datetime.datetime.strptime instead of datetime.strptime
            start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            start_date = first_day_of_month
            end_date = last_day_of_month
    else:
        start_date = first_day_of_month
        end_date = last_day_of_month

    # All data for charts and totals
    incomes = Income.objects.filter(
        user=user,
        date__range=[start_date, end_date],
        amount__gt=0
    ).order_by('-date')
    
    expenses = Expense.objects.filter(
        user=user,
        date__range=[start_date, end_date],
        amount__gt=0
    ).order_by('-date')


    total_income = sum(i.amount for i in incomes)
    total_expense = sum(e.amount for e in expenses)
    balance = total_income - total_expense

    # Budget vs Actual-specific expenses (exclude savings-generated)
    expenses_for_budget_vs_actual = Expense.objects.filter(
        user=request.user, date__range=[start_date, end_date]
    ).exclude(category__isnull=True, subcategory__isnull=True)
    
    budget_items = BudgetItem.objects.filter(user=user, month__range=[start_date, end_date]).order_by('-month')
    summary = defaultdict(lambda: {'budgeted': 0, 'actual': 0})

    total_budget = sum(item.forecasted_amount for item in budget_items)
    total_spent = sum(exp.amount for exp in expenses_for_budget_vs_actual)

    budget_pct = 100
    spent_pct = round((total_spent / total_budget * 100), 1) if total_budget > 0 else 0

    # Category aggregation (for charts) â€” uses all expenses
    category_expenses = defaultdict(float)
    for exp in expenses:
        if exp.subcategory and exp.subcategory.category:
            key = exp.subcategory.category.name
        elif exp.category:
            key = exp.category.name
        else:
            key = 'Uncategorized'
        category_expenses[key] += float(exp.amount)

    # Budgeted amounts
    for item in budget_items:
        key = item.subcategory.name if item.subcategory else item.category.name
        summary[key]['budgeted'] += float(item.forecasted_amount)

    # Actual spending for Budget vs Actual â€” uses filtered expenses
    for exp in expenses_for_budget_vs_actual:
        key = exp.subcategory.name if exp.subcategory else exp.category.name
        summary[key]['actual'] += float(exp.amount)

    # Subcategory aggregation for donut chart â€” uses all expenses
    subcategory_expenses = defaultdict(float)
    for exp in expenses:
        if exp.subcategory:
            key = exp.subcategory.name
        else:
            key = 'Uncategorized'
        subcategory_expenses[key] += float(exp.amount)

    # Build report table
    #report = []
    #for key, data in summary.items():
    #    data['difference'] = data['budgeted'] - data['actual']
    #    report.append({
    #        'item': key,
    #        **data
    #    })

    report = []
    for key, data in summary.items():
        data['difference'] = data['budgeted'] - data['actual']
        
        # Only include if at least one > 0
        if data['budgeted'] > 0 or data['actual'] > 0:
            report.append({
                'item': key,
                **data
            })


    context = {
        'incomes': incomes,
        'expenses': expenses,
        'total_income': total_income,
        'total_expense': total_expense,
        'balance': balance,
        'start_date': start_date.strftime('%Y-%m-%d'),
        'end_date': end_date.strftime('%Y-%m-%d'),
        'report': report,
        'month': f"{start_date.strftime('%B')} â€“ {end_date.strftime('%B %Y')}" if start_date.month != end_date.month else start_date.strftime('%B %Y'),
        'total_budget': total_budget,
        'total_spent': total_spent,
        'budget_pct': budget_pct,
        'spent_pct': spent_pct,
        'category_expenses': category_expenses,
        'category_expenses_json': json.dumps(category_expenses),
        'subcategory_expenses': subcategory_expenses,
        'subcategory_expenses_json': json.dumps(subcategory_expenses),
    }

    return render(request, 'budget/custom_report.html', context)



@login_required
def export_report_excel(request):
    user = request.user

    # Get date range from query params
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    try:
        start_date = datetime.datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        today = date.today()
        start_date = today.replace(day=1)
        end_date = today.replace(day=calendar.monthrange(today.year, today.month)[1])


    incomes = Income.objects.filter(user=user, date__range=[start_date, end_date])
    expenses = Expense.objects.filter(user=user, date__range=[start_date, end_date])
    budget_items = BudgetItem.objects.filter(user=user, month__range=[start_date, end_date])

    # Create workbook
    wb = openpyxl.Workbook()

    # === Incomes Sheet ===
    ws1 = wb.active
    ws1.title = "Income"
    ws1.append(["Date", "Income Category", "Income", "Transaction Method", "Description", "Amount"])
    for income in incomes:
        ws1.append([
            income.date.strftime("%Y-%m-%d"),
            income.category.name if income.category else '',
            income.subcategory.name if income.subcategory else '',
            income.income_source.name if income.income_source else '',
            income.source or '',
            float(income.amount),
        ])

    # === Expenses Sheet ===
    ws2 = wb.create_sheet(title="Expenses")
    ws2.append(["Date", "Expense Category", "Expense", "Transaction Method", "Description", "Amount"])
    for expense in expenses:
        ws2.append([
            expense.date.strftime("%Y-%m-%d"),
            expense.category.name if expense.category else '',
            expense.subcategory.name if expense.subcategory else '',
            expense.source.name if expense.source else '',
            expense.description or '',
            float(expense.amount),
        ])

    # === Budget vs Actual Sheet ===
    ws3 = wb.create_sheet(title="Budget vs Actual")
    summary = defaultdict(lambda: {'budgeted': 0, 'actual': 0})

    # Aggregate budgeted amounts
    for item in budget_items:
        key = item.subcategory.name if item.subcategory else item.category.name
        summary[key]['budgeted'] += float(item.forecasted_amount)

    # Aggregate actual expenses
    for exp in expenses:
        key = exp.subcategory.name if exp.subcategory else exp.category.name
        summary[key]['actual'] += float(exp.amount)

    # Write headers
    ws3.append(["Item", "Budgeted (KES)", "Spent (KES)", "Difference (KES)"])

    for key, data in summary.items():
        difference = data['budgeted'] - data['actual']
        ws3.append([key, data['budgeted'], data['actual'], difference])

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"budget_report_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


############################################################################################ WALLETS #########################################################################################################################################

@login_required
def wallet_overview(request):
    user = request.user

    # all wallets for this user
    wallets = ExpenseSource.objects.filter(Q(user=user) | Q(user__isnull=True))

    # wallet summary data
    wallet_data = []
    for wallet in wallets:
        income_total = Income.objects.filter(user=user, income_source=wallet).aggregate(Sum('amount'))['amount__sum'] or 0
        expense_total = Expense.objects.filter(user=user, source=wallet).aggregate(Sum('amount'))['amount__sum'] or 0

        # Skipping wallets with no income and no expense
        if income_total == 0 and expense_total == 0:
            continue

        balance = income_total - expense_total
        wallet_data.append({
            'name': wallet.name,
            'income_total': income_total,
            'expense_total': expense_total,
            'balance': balance,
        })

    # Combining income and expense into unified transaction list
    income_records = Income.objects.filter(user=user, income_source__isnull=False).values(
        'date', 'amount', 'income_source__name', 'source'
    )
    expense_records = Expense.objects.filter(user=user, source__isnull=False).values(
        'date', 'amount', 'source__name', 'description'
    )

    transactions = []

    for i in income_records:
        transactions.append({
            'type': 'income',
            'date': i['date'],
            'amount': i['amount'],
            'wallet': {'name': i['income_source__name']},
            'description': i['source'] or 'No description'
        })

    for e in expense_records:
        transactions.append({
            'type': 'expense',
            'date': e['date'],
            'amount': e['amount'],
            'wallet': {'name': e['source__name']},
            'description': e['description'] or 'No description'
        })

    # Sort by date (latest first)
    transactions.sort(key=lambda x: x['date'], reverse=True)

    return render(request, 'budget/wallets.html', {
        'wallet_data': wallet_data,
        'transactions': transactions,
    })


def onboarding_view(request): 
    return render(request, 'budget/onboarding_carousel.html')


from .forms import ProfileForm

@login_required
def account_settings(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Account updated successfully.")
            return redirect('my_dashboard')  # Or reload the same page to show modal again
    else:
        form = ProfileForm(instance=request.user)
    return render(request, 'budget/my_dashboard.html', {'form': form})


from django.http import JsonResponse

@login_required
def update_profile(request):
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=request.user)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            if form.is_valid():
                form.save()
                return JsonResponse({'success': True, 'message': 'Your profile was updated successfully.'})
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                }, status=400)
        else:
            # fallback if form posted without AJAX
            if form.is_valid():
                form.save()
                messages.success(request, "Your profile was updated successfully.")
            else:
                messages.warning(request, "Please correct the errors below.")
            return redirect(request.META.get('HTTP_REFERER', 'home'))


############################################################################################ LOANS #########################################################################################################################################

# @login_required
# def loan_list(request):
#     loans = LoanAccount.objects.filter(user=request.user)
#     return render(request, 'budget/loan_list.html', {'loans': loans})

@login_required
def loan_list(request):
    loans = LoanAccount.objects.filter(user=request.user)
    
    # Get savings-generated transactions for each goal
    loan_data = []
    for loan in loans:
        transactions = Expense.objects.filter(
            user=request.user,
            subcategory__isnull=False,          # make sure subcategory exists
            subcategory__name=loan.name,        # match loan name
            category__is_loan_generated=True,
            amount__gt=0                        # exclude placeholder
        ).order_by('-date')


        loan_data.append({
            'loan': loan,
            'transactions': transactions
        })
    
    return render(request, 'budget/loan_list.html', {'loans': loans, 'loan_data': loan_data})


@login_required
def create_loan(request):
    # 🔹 Ensure "Debt/Loan Repayment" category exists
    category, _ = ExpenseCategory.objects.get_or_create(
        name="Debt/Loan Repayment",
        user=request.user,
        defaults={'is_loan_generated': True}
    )

    if request.method == 'POST':
        form = LoanAccountForm(request.POST)
        if form.is_valid():
            loan = form.save(commit=False)
            loan.user = request.user

            # 🔹 Prevent duplicate loan names
            duplicate = LoanAccount.objects.filter(
                user=request.user,
                name__iexact=loan.name
            ).exclude(pk=loan.pk).exists()
            if duplicate:
                messages.warning(request, f"A loan named '{loan.name}' already exists. Please see your loans below.")
                return redirect('loan_list')

            loan.save()

            # 🔹 Create a subcategory tied to this loan
            subcategory, _ = ExpenseSubCategory.objects.get_or_create(
                category=category,
                name=loan.name,
                user=request.user,
            )

            # 🔹 Always add a zero-amount expense so it appears in budgets
            Expense.objects.create(
                user=request.user,
                category=category,
                subcategory=subcategory,
                description=f"Loan account created: {loan.name}",
                amount=0,
                date=timezone.now().date()
            )

            messages.success(request, f"Loan '{loan.name}' created successfully.")
            return redirect('loan_detail', loan_id=loan.id)
    else:
        form = LoanAccountForm()

    return render(request, 'budget/create_loan.html', {'form': form})

FREQ_MAP = {
    'monthly': 12,
    'weekly': 52,
    'yearly': 1
}

@login_required
def loan_detail(request, loan_id):
    loan = get_object_or_404(LoanAccount, id=loan_id, user=request.user)
    payments_qs = loan.payments.all().order_by('date')

    # Convert string frequency to numeric periods per year
    freq_value = FREQ_MAP.get(loan.payment_frequency, 12)

    # Build actual schedule
    schedule, total_interest, payoff_periods = loan_amortization_schedule(
        principal=loan.principal,
        annual_interest_rate=loan.annual_interest_rate,
        term_periods=loan.term_periods,
        payments_per_year=freq_value
    )

    # Extra payment simulation inputs
    try:
        extra_per_period = Decimal(request.GET.get('extra_per_period', 0) or 0)
    except:
        extra_per_period = Decimal('0')

    try:
        one_time_extra = Decimal(request.GET.get('one_time_extra', 0) or 0)
    except:
        one_time_extra = Decimal('0')

    try:
        start_extra_period = int(request.GET.get('start_extra_period', 0) or 0)
    except:
        start_extra_period = 0

    # Simulation schedule
    sim_schedule, sim_interest, sim_periods = loan_amortization_schedule(
        principal=loan.principal,
        annual_interest_rate=loan.annual_interest_rate,
        term_periods=loan.term_periods,
        payments_per_year=freq_value,
        extra_payment=extra_per_period,
        one_time_extra=one_time_extra,
        start_extra_period=start_extra_period
    )

    # Prepare payments for chart
    payments_list = [
        {"period": i + 1, "amount": float(p.amount)}
        for i, p in enumerate(payments_qs)
    ]


    # Calculate interest saved and time saved
    interest_saved = total_interest - sim_interest
    time_saved_periods = payoff_periods - sim_periods


    # Convert periods to years & months
    if time_saved_periods > 0:
        years_saved = time_saved_periods // freq_value
        months_saved = (time_saved_periods % freq_value) * (12 // freq_value)
        if years_saved > 0 and months_saved > 0:
            time_saved_str = f"{years_saved} year{'s' if years_saved > 1 else ''} and {months_saved} month{'s' if months_saved > 1 else ''}"
        elif years_saved > 0:
            time_saved_str = f"{years_saved} year{'s' if years_saved > 1 else ''}"
        else:
            time_saved_str = f"{months_saved} month{'s' if months_saved > 1 else ''}"
    else:
        time_saved_str = "No time saved"



    # Balances
    amount_paid_so_far = sum(p.amount for p in payments_qs)
    current_balance = loan.principal - Decimal(amount_paid_so_far)

    if sim_schedule:
        if len(sim_schedule) > len(payments_qs):
            projected_balance_now = sim_schedule[len(payments_qs)]['balance']
        else:
            projected_balance_now = 0
    else:
        projected_balance_now = 0

    payments_made = LoanPayment.objects.filter(loan=loan).count()
    remaining_payments = loan.term_periods - payments_made


    # Send to template
    context = {
        'loan': loan,
        'payments': json.dumps(payments_list, cls=DjangoJSONEncoder),
        'schedule': schedule,
        'total_interest': total_interest,
        'payoff_periods': payoff_periods,
        'sim_schedule': sim_schedule,
        'sim_interest': sim_interest,
        'sim_periods': sim_periods,
        'extra_per_period': extra_per_period,
        'one_time_extra': one_time_extra,
        'start_extra_period': start_extra_period,
        'interest_saved': interest_saved,
        'time_saved_str': time_saved_str,
        'projected_balance_now': projected_balance_now,
        'current_balance': current_balance,
        'remaining_payments': remaining_payments,
    }
    return render(request, 'budget/loan_detail.html', context)


@login_required
def add_loan_payment(request, loan_id):
    loan = get_object_or_404(LoanAccount, id=loan_id, user=request.user)

    if request.method == 'POST':
        form = LoanPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.loan = loan
            payment.save()

            # Save as Expense
            category, _ = ExpenseCategory.objects.get_or_create(
                name="Debt/Loan Repayment",
                user=request.user
            )
            subcategory, _ = ExpenseSubCategory.objects.get_or_create(
                name=loan.name,
                category=category,
                user=request.user
            )
            Expense.objects.create(
                user=request.user,
                category=category,
                subcategory=subcategory,
                description=f"Loan repayment for {loan.name}",
                amount=payment.amount,
                date=payment.date,
                source=get_object_or_404(ExpenseSource, id=request.POST.get('source_id')),
            )

            messages.success(request, f"Expense of {payment.amount} recorded for {loan.name}.")
            return redirect('loan_detail', loan_id=loan.id)
    else:
        form = LoanPaymentForm()

    # ✅ Exclude hidden wallets
    hidden_wallet_ids = UserHiddenWallet.objects.filter(
        user=request.user, hidden=True
    ).values_list('wallet_id', flat=True)
    
    # ✅ Wallets: user + global, non-system, and not hidden
    wallets = ExpenseSource.objects.filter(
        (Q(user=request.user) | Q(user__isnull=True))
    ).exclude(
        id__in=hidden_wallet_ids
    ).order_by("name")


    return render(request, 'budget/add_loan_payment.html', {'form': form, 'loan': loan, 'wallets': wallets})

@login_required
def loan_delete(request, pk):
    loan = get_object_or_404(LoanAccount, pk=pk, user=request.user)

    if request.method == "POST":
        loan.delete()
        messages.success(request, "Loan account deleted successfully.")
        return redirect('loan_list')  # Change to your loans list page

    messages.error(request, "Invalid request.")
    return redirect('loan_detail', pk=pk)


@login_required
def delete_selected_incomes(request):
    if request.method == "POST":
        ids = request.POST.getlist("selected_incomes")
        Income.objects.filter(user=request.user, id__in=ids).delete()
    return redirect("dashboard")


@login_required
def delete_selected_expenses(request):
    if request.method == "POST":
        ids = request.POST.getlist("selected_expenses")
        Expense.objects.filter(user=request.user, id__in=ids).delete()
    return redirect("dashboard")

# views.py
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils.timezone import now

@login_required
def toggle_favourite_income(request, income_id):
    income = get_object_or_404(Income, id=income_id, user=request.user)
    income.is_favourite = not income.is_favourite
    income.save()
    messages.success(request, f"Income {'bookmarked' if income.is_favourite else 'removed from bookmarked'}.")
    return redirect('dashboard')

@login_required
def duplicate_income(request, income_id):
    income = get_object_or_404(Income, id=income_id, user=request.user)
    new_income = Income.objects.create(
        user=request.user,
        category=income.category,
        subcategory=income.subcategory,
        income_source=income.income_source,
        amount=income.amount,
        date=now().date(),  # use today’s date by default
        source=income.source
    )
    messages.success(request, "Income duplicated successfully.")
    return redirect('dashboard')


@login_required
def toggle_favourite_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    expense.is_favourite = not expense.is_favourite
    expense.save()
    messages.success(request, f"Expensee {'added to' if expense.is_favourite else 'removed from'} bookmarked.")
    return redirect('dashboard')

@login_required
def duplicate_expense(request, expense_id):
    expense = get_object_or_404(Expense, id=expense_id, user=request.user)
    new_expense = Expense.objects.create(
        user=request.user,
        category=expense.category,
        subcategory=expense.subcategory,
        description=expense.description,
        amount=expense.amount,
        date=now().date(),  # use today’s date by default
        source=expense.source
    )
    messages.success(request, "Expense duplicated successfully.")
    return redirect('dashboard')
    

from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.utils.timezone import now
from .models import BudgetItem


@login_required
def toggle_favourite_budget(request, budget_id):
    budget = get_object_or_404(BudgetItem, id=budget_id, user=request.user)
    budget.is_favourite = not budget.is_favourite
    budget.save()
    messages.success(request, f"Budget {'bookmarked' if budget.is_favourite else 'removed from bookmarked'}.")
    return redirect('all_budget_vs_actual')

@login_required
def duplicate_budget(request, budget_id):
    budget = get_object_or_404(BudgetItem, id=budget_id, user=request.user)

    current_month = now().date().replace(day=1)  # normalize to first of the month

    # Check if same subcategory exists this month
    exists = BudgetItem.objects.filter(
        user=request.user,
        month=current_month,
        subcategory=budget.subcategory
    ).exists()

    if exists:
        messages.warning(
            request,
            f"A budget for '{budget.subcategory}' already exists for {current_month.strftime('%B %Y')}."
        )
    else:
        BudgetItem.objects.create(
            user=request.user,
            month=current_month,
            category=budget.category,
            subcategory=budget.subcategory,
            forecasted_amount=budget.forecasted_amount,
        )
        messages.success(request, "Budget duplicated successfully.")

    return redirect('all_budget_vs_actual')


from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Income, ExpenseSource
import datetime

@login_required
def add_starting_balance(request):
    user = request.user

    # --- Filter wallets: user + global, exclude system + hidden ---
    hidden_wallet_ids = UserHiddenWallet.objects.filter(user=user, hidden=True).values_list('wallet_id', flat=True)
    wallets = ExpenseSource.objects.filter(
        Q(user=user) | Q(user__isnull=True)
    ).exclude(
        Q(id__in=hidden_wallet_ids)
    ).order_by('name')

    if request.method == 'POST':
        total_forms = int(request.POST.get('form-TOTAL_FORMS', 1))
        added = 0
        for i in range(total_forms):
            wallet_id = request.POST.get(f'wallet_{i}')
            amount = request.POST.get(f'amount_{i}')
            if wallet_id and amount:
                Income.objects.create(
                    user=user,
                    income_source_id=wallet_id,
                    amount=amount,
                    date=datetime.date.today(),
                    category=None,
                    subcategory=None,
                    source='Uncategorized: Added to Wallet'
                )
                added += 1

        if added > 0:
            messages.success(request, f'Successfully added to wallet{"s" if added > 1 else ""}.')

        return redirect('wallet_overview')

    return render(request, 'budget/add_starting_balance.html', {'wallets': wallets})


    
from datetime import date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import ExpenseCategory, ExpenseSubCategory, BudgetItem
from .utils import get_user_subcategories  # helper that excludes hidden ones

@login_required
def add_starting_budget(request):
    user = request.user
    current_month = date.today().replace(day=1)

    # ✅ Categories (excluding savings/loan-generated)
    categories = ExpenseCategory.objects.filter(
        (Q(user=user) | Q(user__isnull=True)) &
        (Q(is_savings_generated=False) | Q(is_savings_generated__isnull=True)) &
        (Q(is_loan_generated=False) | Q(is_loan_generated__isnull=True))
    )

    # ✅ Subcategories (excluding hidden ones)
    subcategories = get_user_subcategories(user).filter(
        Q(category__user=user) | Q(category__user__isnull=True)
    )

    if request.method == 'POST':
        month = request.POST.get('month')
        total_forms = int(request.POST.get('form-TOTAL_FORMS', 1))
        added, skipped = 0, 0

        for i in range(total_forms):
            subcat_id = request.POST.get(f'subcategory_{i}')
            amount = request.POST.get(f'amount_{i}')

            if subcat_id and amount:
                try:
                    subcat = ExpenseSubCategory.objects.get(id=subcat_id)
                except ExpenseSubCategory.DoesNotExist:
                    continue  # skip invalid or hidden IDs

                exists = BudgetItem.objects.filter(
                    user=user,
                    month=month,
                    subcategory=subcat
                ).exists()

                if exists:
                    skipped += 1
                    continue

                BudgetItem.objects.create(
                    user=user,
                    month=month,
                    category=subcat.category,
                    subcategory=subcat,
                    forecasted_amount=amount
                )
                added += 1

        # ✅ Messages
        if added:
            messages.success(
                request,
                f"{added} budget{'s' if added > 1 else ''} saved successfully."
            )
        if skipped:
            messages.warning(
                request,
                f"{skipped} budget{'s' if skipped > 1 else ''} item{'s' if skipped > 1 else ''} already exist."
            )

        return redirect('all_budget_vs_actual')

    return render(request, 'budget/add_starting_budget.html', {
        'current_month': current_month,
        'categories': categories,
        'subcategories': subcategories,
    })

    
# views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import InvestmentAccount, InvestmentTransaction, PriceHistory
from .utils import update_prices
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import InvestmentAccountForm
from django.db.models import Sum, Q, F
from django.shortcuts import render
import json
from datetime import date, timedelta
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from decimal import Decimal
from collections import defaultdict
from django.core.exceptions import ValidationError
from .forms import InvestmentAccountForm
from .forms import InvestmentTransactionForm



@login_required
def portfolio_dashboard(request):
    accounts = InvestmentAccount.objects.filter(user=request.user)
    portfolio_data = []

    total_invested = Decimal('0')
    total_value = Decimal('0')
    total_dividends = Decimal('0')
    total_interest = Decimal('0')
    total_rental_income = Decimal('0')
    total_expenses = Decimal('0')

    gain_loss_data = []
    gain_loss_labels = []

    # --- Per-asset-type totals ---
    totals_by_type = defaultdict(lambda: {
        "invested": Decimal('0'),
        "current_value": Decimal('0'),
        "dividends": Decimal('0'),
        "interest": Decimal('0'),
        "rental_income": Decimal('0'),
        "expenses": Decimal('0'),
        "gain_loss": Decimal('0'),
        "realized_gain_loss": Decimal('0'),
        "unrealized_gain_loss": Decimal('0'),
        "total_return": Decimal('0'),
    })

    for acc in accounts:
        txs = InvestmentTransaction.objects.filter(account=acc)

        # --- Transaction breakdown ---
        bought_units = txs.filter(action="BUY").aggregate(total=Sum("units"))["total"] or Decimal('0')
        sold_units = txs.filter(action="SELL").aggregate(total=Sum("units"))["total"] or Decimal('0')
        units = (bought_units or Decimal("0")) - (sold_units or Decimal("0")) + (acc.units_held or Decimal("0"))

        invested = txs.filter(action="BUY").aggregate(total=Sum("amount"))["total"] or Decimal('0')
        dividends = txs.filter(action="DIVIDEND").aggregate(total=Sum("amount"))["total"] or Decimal('0')
        interest = txs.filter(action="INTEREST").aggregate(total=Sum("amount"))["total"] or Decimal('0')
        rental_income = txs.filter(action="RENTAL").aggregate(total=Sum("amount"))["total"] or Decimal('0')
        expenses = txs.filter(action="EXPENSE").aggregate(total=Sum("amount"))["total"] or Decimal('0')

        # --- Latest price lookup ---
        latest_price_entry = PriceHistory.objects.filter(account=acc).order_by('-date').first()
        if latest_price_entry:
            current_price = Decimal(latest_price_entry.price)
            currency = latest_price_entry.currency or "USD"
        else:

            # Get last transaction safely
            last_tx = txs.last()  # returns None if no tx
            current_price = Decimal(last_tx.price_per_unit or 0) if last_tx else Decimal(0)

            currency = "USD" if acc.asset_type == "crypto" else "KES"

        # --- Asset-specific calculations ---
        current_value = units * current_price if units else Decimal('0')
        gain_loss = Decimal('0')
        projected_interest = Decimal('0')
        accrued_interest = Decimal('0')
        realized_gain_loss = Decimal('0')
        pct_gain_loss = Decimal('0')
        roi_pct = Decimal('0')
        total_return = Decimal('0')
        dividend_yield = Decimal('0')
        income = Decimal('0')
        capital_shares = Decimal('0')
        contributions = Decimal('0')
        all_income = Decimal('0')
        all_expense = Decimal('0')
        expected_interest = Decimal('0')
        coupon_payment = Decimal('0')


        if acc.asset_type in ["stock", "crypto", "etf"]:
            # Initial units from account setup
            units = acc.units_held or Decimal("0")
            units = units + (bought_units or Decimal("0")) - (sold_units or Decimal("0"))

            # Cost basis (initial + net buys - sells)
            invested = (txs.filter(action="BUY").aggregate(total=Sum("amount"))["total"] or Decimal("0"))
            invested -= (txs.filter(action="SELL").aggregate(total=Sum("amount"))["total"] or Decimal("0"))
            invested += (acc.units_held or Decimal("0")) * (acc.face_value or Decimal("0"))  # opening cost
            invested = max(invested, 0)


            # Current value = units × current market price
            current_value = units * (current_price or Decimal("0"))

            realized_gain_loss = Decimal("0")
            unrealized_gain_loss = Decimal("0")
            lot_remaining_units = []

            # Track sells against buys (FIFO)
            sold_units_remaining = sold_units
            buy_lots = list(txs.filter(action="BUY").order_by("date"))

            for lot in buy_lots:
                lot_units = lot.units
                lot_price = lot.price_per_unit

                if sold_units_remaining > 0:
                    units_sold_from_lot = min(lot_units, sold_units_remaining)
                    realized_gain_loss += (current_price - lot_price) * units_sold_from_lot
                    lot_units -= units_sold_from_lot
                    sold_units_remaining -= units_sold_from_lot

                if lot_units > 0:
                    lot_remaining_units.append((lot_units, lot_price))

            # Unrealized = remaining units × (market - cost)
            for units_left, price in lot_remaining_units:
                unrealized_gain_loss += (current_price - price) * units_left

            # Add dividends if any
            dividends = txs.filter(action="DIVIDEND").aggregate(total=Sum("amount"))["total"] or Decimal("0")

            gain_loss = (realized_gain_loss + unrealized_gain_loss)
            total_return = gain_loss + dividends
            pct_gain_loss = (total_return / invested * 100) if invested else Decimal("0")
            dividend_yield = (dividends / current_value * 100) if current_value else Decimal("0")


        elif acc.asset_type == "bond":
            # Current holdings
            initial_units = acc.units_held or Decimal("0")
            units = initial_units + (bought_units or Decimal("0")) - (sold_units or Decimal("0"))

            face_value = acc.face_value or Decimal("0")
            coupon_rate = acc.coupon_rate or Decimal("0")

            # Current market value
            current_value = units * face_value

            # Invested amount: initial units at face value + net transactions
            initial_invested = initial_units * face_value
            tx_buys = txs.filter(action="BUY").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            tx_sells = txs.filter(action="SELL").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            invested = initial_invested + tx_buys - tx_sells
            invested = max(invested, 0)

            # Default interest values
            accrued_interest = Decimal("0")
            projected_interest = Decimal("0")
            expected_interest = Decimal("0")

            # Accrued interest calculation
            if acc.issue_date:
                days_since_issue = (date.today() - acc.issue_date).days
                annual_coupon = units * face_value * (coupon_rate / 100)

                # Coupon frequency
                if acc.interest_frequency == "semiannual":
                    period_days = 365 / 2
                    coupon_payment = annual_coupon / 2
                    periods_per_year = 2
                elif acc.interest_frequency == "quarterly":
                    period_days = 365 / 4
                    coupon_payment = annual_coupon / 4
                    periods_per_year = 4
                elif acc.interest_frequency == "monthly":
                    period_days = 365 / 12
                    coupon_payment = annual_coupon / 12
                    periods_per_year = 12
                else:  # annual
                    period_days = 365
                    coupon_payment = annual_coupon
                    periods_per_year = 1

                # Completed periods + partial period for accrued interest
                periods_completed = int(days_since_issue // period_days)
                days_into_period = days_since_issue % period_days
                accrued_interest = (periods_completed * coupon_payment) + (coupon_payment * Decimal(days_into_period / period_days))

            # Expected and projected interest until maturity
            if acc.issue_date and acc.maturity_date:
                total_days = (acc.maturity_date - acc.issue_date).days
                total_years = Decimal(total_days) / Decimal(365)
                total_periods = int(total_years * periods_per_year)
                expected_interest = coupon_payment * total_periods

                # Projected interest = expected - accrued
                projected_interest = expected_interest - accrued_interest

            # Gains / ROI
            gain_loss = (current_value + accrued_interest) - invested
            total_return = gain_loss + projected_interest
            roi_pct = (gain_loss / invested * 100) if invested else Decimal("0")



        elif acc.asset_type == "mmf":
            yield_rate = acc.yield_rate or Decimal('0')
            daily_rate = yield_rate / Decimal("100") / Decimal("365")

            # Manual starting balance (user-entered)
            base_value = current_price or Decimal("0")
            base_date = acc.base_date or date.today()  # fallback = today

            # Transactions
            txs = txs.order_by("date")
            first_tx_date = txs.aggregate(Min("date"))["date__min"]

            # Starting point = whichever comes first: base_date or first transaction
            if first_tx_date:
                start_date = min(base_date, first_tx_date)
            else:
                start_date = base_date

            balance = base_value
            invested_total = base_value
            last_date = start_date

            # Grow base_value from base_date up to the first transaction
            if first_tx_date and base_date < first_tx_date:
                days = (first_tx_date - base_date).days
                if days > 0:
                    balance *= (Decimal("1") + daily_rate) ** days
                last_date = first_tx_date

            # Apply each transaction with compounding in between
            for tx in txs:
                days = (tx.date - last_date).days
                if days > 0:
                    balance *= (Decimal("1") + daily_rate) ** days
                last_date = tx.date

                if tx.action == "DEPOSIT":
                    balance += tx.amount
                    invested_total += tx.amount
                elif tx.action == "WITHDRAW":
                    balance -= tx.amount
                    invested_total -= tx.amount  # treat withdrawals as reducing invested

            # Compound from last transaction to today
            days = (date.today() - last_date).days
            if days > 0:
                balance *= (Decimal("1") + daily_rate) ** days

            current_value = balance
            gain_loss = current_value - invested_total
            pct_gain_loss = (gain_loss / invested_total * 100) if invested_total else Decimal("0")
            invested = invested_total
            invested = max(invested, 0)


        elif acc.asset_type == "real_estate":
            base_value = current_price or Decimal("0")
            total_deposit = txs.filter(action="DEPOSIT").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            total_withdraw = txs.filter(action="WITHDRAW").aggregate(total=Sum("amount"))["total"] or Decimal("0")

            # principal-like contribution
            net_invested_txns = total_deposit - total_withdraw
            invested = base_value + net_invested_txns
            invested = max(invested, 0)

            # actual value now
            rental_income = txs.filter(action="RENTAL").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            expenses = txs.filter(action="EXPENSE").aggregate(total=Sum("amount"))["total"] or Decimal("0")

            all_income = acc.income_so_far + rental_income
            all_expense = acc.expenses_so_far + expenses           

            current_value = invested + all_income - all_expense
            gain_loss = current_value - invested
            roi_pct = (gain_loss / invested * 100) if invested else Decimal("0")

        elif acc.asset_type == "other":
            base_value = current_price or Decimal("0")
            total_deposit = txs.filter(action="DEPOSIT").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            total_withdraw = txs.filter(action="WITHDRAW").aggregate(total=Sum("amount"))["total"] or Decimal("0")

            # principal-like contribution
            net_invested_txns = total_deposit - total_withdraw
            invested = base_value + net_invested_txns
            invested = max(invested, 0)

            # flow-based PnL
            income = txs.filter(action="INTEREST").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            expenses = txs.filter(action="EXPENSE").aggregate(total=Sum("amount"))["total"] or Decimal("0")

            all_income = acc.income_so_far + income
            all_expense = acc.expenses_so_far + expenses 

            current_value = invested + all_income - all_expense
            gain_loss = current_value - invested
            roi_pct = (gain_loss / invested * 100) if invested else Decimal("0")


        elif acc.asset_type == "sacco":
            # --- Invested: Capital shares + Regular contributions ---
            capital_shares = txs.filter(action="CAPITAL_SHARE").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            contributions = txs.filter(action="CONTRIBUTION").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            invested = acc.sacco_capital_share + capital_shares + contributions + acc.sacco_total_contributions
            invested = max(invested, 0)
            dividends = txs.filter(action="DIVIDEND").aggregate(total=Sum("amount"))["total"] or Decimal("0")
            income = dividends + acc.sacco_dividends_earned
            current_value = invested + income
            gain_loss = current_value - invested
            roi_pct = (gain_loss / invested * 100) if invested else Decimal("0")

        else:
            current_value = units * current_price if units else Decimal('0')
            gain_loss = (current_value + dividends) - invested

        # --- Collect portfolio data ---
        portfolio_data.append({
            "account": acc,
            "symbol": acc.symbol or "N/A",
            "units": units,
            "invested": invested,
            "dividends": dividends,
            "interest": interest,
            "rental_income": rental_income,
            "expenses": expenses,
            "current_price": current_price,
            "currency": currency,
            "current_value": current_value,
            "gain_loss": gain_loss,
            "coupon_payment": coupon_payment,
            "projected_interest": projected_interest,
            "expected_interest": expected_interest,
            "accrued_interest": accrued_interest,
            "realized_gain_loss": realized_gain_loss,
            "pct_gain_loss": pct_gain_loss,
            "roi_pct": roi_pct,
            "total_return": total_return,
            "dividend_yield": dividend_yield,
            "income": income,
            "capital_shares": capital_shares,
            "contributions": contributions,
            "all_income": all_income,
            "all_expense": all_expense,

        })

        # --- For charts ---
        gain_loss_labels.append(acc.name)
        gain_loss_data.append(float(gain_loss))

        # --- Update global totals ---
        total_invested += invested
        total_value += current_value
        total_dividends += dividends
        total_interest += interest
        total_rental_income += rental_income
        total_expenses += expenses

        # --- Update per-type totals ---
        asset_type = "stocks" if acc.asset_type in ["stock", "crypto", "etf"] else acc.asset_type

        totals_by_type[asset_type]["invested"] += invested
        totals_by_type[asset_type]["current_value"] += current_value
        totals_by_type[asset_type]["dividends"] += dividends
        totals_by_type[asset_type]["interest"] += interest  # 0 for SACCO
        totals_by_type[asset_type]["rental_income"] += rental_income  # 0 for SACCO
        totals_by_type[asset_type]["expenses"] += expenses  # 0 for SACCO
        totals_by_type[asset_type]["gain_loss"] += gain_loss
        totals_by_type[asset_type]["realized_gain_loss"] += realized_gain_loss  # 0 for SACCO
        totals_by_type[asset_type]["total_return"] += gain_loss + income



        if asset_type == "bond":
            totals_by_type[asset_type]["accrued_interest"] = totals_by_type[asset_type].get("accrued_interest", Decimal("0")) + accrued_interest
            totals_by_type[asset_type]["projected_interest"] = totals_by_type[asset_type].get("projected_interest", Decimal("0")) + projected_interest


    # --- Historical data for charts ---

    # 1️⃣ Determine earliest possible date across all accounts
    earliest_dates = []
    for acc in accounts:
        tx_earliest = txs.aggregate(Min('date'))['date__min']
        price_earliest = PriceHistory.objects.filter(account=acc).aggregate(Min('date'))['date__min']
        earliest = min(filter(None, [tx_earliest, price_earliest]), default=date.today())
        earliest_dates.append(earliest)

    first_possible_date = min(earliest_dates) if earliest_dates else date.today()

    # 2️⃣ Build preliminary full date range
    num_days = (date.today() - first_possible_date).days + 1
    full_dates = [first_possible_date + timedelta(days=i) for i in range(num_days)]

    # 3️⃣ Compute account growth and prices
    account_growth = {}
    account_prices = {}
    for acc in accounts:
        values = []
        prices = []
        for day in full_dates:
            units_day = InvestmentTransaction.objects.filter(account=acc, date__lte=day).aggregate(total_units=Sum('units'))['total_units'] or Decimal('0')
            price_entry = PriceHistory.objects.filter(account=acc, date__lte=day).order_by('-date').first()
            price = Decimal(price_entry.price) if price_entry else Decimal('0')
            values.append(float(units_day * price))
            prices.append(float(price))
        account_growth[acc.name] = values
        account_prices[acc.name] = prices

    # 4️⃣ Trim the timeline to start at first non-zero total portfolio value
    for i, day in enumerate(full_dates):
        total_value = sum(account_growth[acc.name][i] for acc in accounts)
        if total_value > 0:
            start_index = i
            break
    else:
        start_index = 0  # fallback if all zeros

    growth_dates = [(full_dates[i]).strftime("%Y-%m-%d") for i in range(start_index, num_days)]
    historical_dates = growth_dates.copy()

    # Trim account data to match new start
    for acc in accounts:
        account_growth[acc.name] = account_growth[acc.name][start_index:]
        account_prices[acc.name] = account_prices[acc.name][start_index:]


    # --- Stock totals for summary row ---
    stock_totals = totals_by_type.get("stocks", {})

    total_stock_invested = stock_totals.get("invested", Decimal("0"))
    total_stock_value = stock_totals.get("current_value", Decimal("0"))
    total_stock_gain_loss = stock_totals.get("gain_loss", Decimal("0"))
    total_realized_gain_loss = stock_totals.get("realized_gain_loss", Decimal("0"))
    total_dividends = stock_totals.get("dividends", Decimal("0"))
    total_return = stock_totals.get("total_return", Decimal("0"))

    # Derived metrics
    total_stock_pct_gain = (total_return / total_stock_invested * 100) if total_stock_invested else Decimal("0")
    dividend_yield = (total_dividends / total_stock_value * 100) if total_stock_value else Decimal("0")

    # --- Bond totals for summary row ---
    bond_totals = totals_by_type.get("bond", {})

    total_bond_invested = bond_totals.get("invested", Decimal("0"))
    total_bond_value = bond_totals.get("current_value", Decimal("0"))
    total_bond_gain_loss = bond_totals.get("gain_loss", Decimal("0"))  # Unrealized price difference
    total_bond_realized_gain_loss = bond_totals.get("realized_gain_loss", Decimal("0"))
    total_bond_interest = bond_totals.get("interest", Decimal("0"))  # Coupon payments received
    total_bond_accrued_interest = bond_totals.get("accrued_interest", Decimal("0"))  # earned but not yet paid
    total_bond_projected_interest = bond_totals.get("projected_interest", Decimal("0"))  # expected coupons till maturity
    total_bond_return = bond_totals.get("total_return", Decimal("0"))

    # Derived metrics
    total_bond_pct_gain = (
        (total_bond_return / total_bond_invested * 100)
        if total_bond_invested else Decimal("0")
    )
    bond_yield = (
        (total_bond_interest / total_bond_value * 100)
        if total_bond_value else Decimal("0")
    )

    total_mmf_invested = sum(p["invested"] for p in portfolio_data if p["account"].asset_type == "mmf")
    total_mmf_value = sum(p["current_value"] for p in portfolio_data if p["account"].asset_type == "mmf")
    total_mmf_accrued = sum(p.get("accrued_interest", 0) for p in portfolio_data if p["account"].asset_type == "mmf")
    total_mmf_gain_loss = sum(p["gain_loss"] for p in portfolio_data if p["account"].asset_type == "mmf")


    # Real Estate totals
    total_real_estate_invested = sum(p["invested"] for p in portfolio_data if p["account"].asset_type == "real_estate")
    total_real_estate_value = sum(p["current_value"] for p in portfolio_data if p["account"].asset_type == "real_estate")
    total_real_estate_rental_income = sum(p.get("rental_income", 0) for p in portfolio_data if p["account"].asset_type == "real_estate")
    total_real_estate_expenses = sum(p.get("expenses", 0) for p in portfolio_data if p["account"].asset_type == "real_estate")
    total_real_estate_gain_loss = sum(p["gain_loss"] for p in portfolio_data if p["account"].asset_type == "real_estate")

    # ROI summary (handle divide by zero safely)
    total_real_estate_roi = (
        (total_real_estate_gain_loss / total_real_estate_invested * 100)
        if total_real_estate_invested else 0
    )

    # Other assets totals
    total_other_invested = sum(p["invested"] for p in portfolio_data if p["account"].asset_type == "other")
    total_other_value = sum(p["current_value"] for p in portfolio_data if p["account"].asset_type == "other")
    total_other_income = sum(p.get("income", 0) for p in portfolio_data if p["account"].asset_type == "other")
    total_other_expenses = sum(p.get("expenses", 0) for p in portfolio_data if p["account"].asset_type == "other")
    total_other_gain_loss = sum(p["gain_loss"] for p in portfolio_data if p["account"].asset_type == "other")

    total_other_roi = (
        (total_other_gain_loss / total_other_invested * 100)
        if total_other_invested else 0
    )


    # SACCO totals
    total_sacco_invested = sum(p["invested"] for p in portfolio_data if p["account"].asset_type == "sacco")
    total_sacco_value = sum(p["current_value"] for p in portfolio_data if p["account"].asset_type == "sacco")
    total_sacco_dividends = sum(p.get("income", 0) for p in portfolio_data if p["account"].asset_type == "sacco")
    total_sacco_gain_loss = sum(p["gain_loss"] for p in portfolio_data if p["account"].asset_type == "sacco")

    total_sacco_roi = (
        (total_sacco_gain_loss / total_sacco_invested * 100)
        if total_sacco_invested else Decimal("0")
    )

    total_current_value = (
        total_stock_value
        + total_bond_value
        + total_mmf_value
        + total_real_estate_value
        + total_other_value
        + total_sacco_value
    )


    total_gain_loss = (
        Decimal(total_current_value or 0)
        + Decimal(total_dividends or 0)
        + Decimal(total_interest or 0)
        + Decimal(total_rental_income or 0)
        - Decimal(total_expenses or 0)
    ) - Decimal(total_invested or 0)


    total_pct_gain = (total_gain_loss / Decimal(total_invested or 1)) * 100

    existing_asset_types = set(acc.get("account").asset_type for acc in portfolio_data if acc.get("account"))


    context = {
        "portfolio_data": portfolio_data,
        "total_invested": total_invested,
        "total_value": total_value,
        "total_dividends": total_dividends,
        "total_interest": total_interest,
        "total_rental_income": total_rental_income,
        "total_expenses": total_expenses,

        "total_gain_loss": total_gain_loss,
        "growth_dates_json": json.dumps(growth_dates),
        "account_growth_json": json.dumps(account_growth),
        "gain_loss_labels_json": json.dumps(gain_loss_labels),
        "gain_loss_data_json": json.dumps(gain_loss_data),
        "historical_dates_json": json.dumps(historical_dates),
        "account_prices_json": json.dumps(account_prices),
        "totals_by_type": dict(totals_by_type),

        # --- Stock summary row context ---
        "total_stock_invested": total_stock_invested,
        "total_stock_value": total_stock_value,
        "total_stock_gain_loss": total_stock_gain_loss,
        "total_realized_gain_loss": total_realized_gain_loss,
        "total_stock_pct_gain": total_stock_pct_gain,
        "total_dividends": total_dividends,
        "dividend_yield": dividend_yield,
        "total_return": total_return,

        "total_bond_invested": total_bond_invested,
        "total_bond_value": total_bond_value,
        "total_bond_gain_loss": total_bond_gain_loss,
        "total_bond_realized_gain_loss": total_bond_realized_gain_loss,
        "total_bond_interest": total_bond_interest,
        "total_bond_accrued_interest": total_bond_accrued_interest,
        "total_bond_projected_interest": total_bond_projected_interest,
        "total_bond_pct_gain": total_bond_pct_gain,
        "bond_yield": bond_yield,
        "total_bond_return": total_bond_return,

        "total_mmf_invested": total_mmf_invested,
        "total_mmf_value": total_mmf_value,
        "total_mmf_accrued": total_mmf_accrued,
        "total_mmf_gain_loss": total_mmf_gain_loss,

        "total_real_estate_invested": total_real_estate_invested,
        "total_real_estate_value": total_real_estate_value,
        "total_real_estate_rental_income": total_real_estate_rental_income,
        "total_real_estate_expenses": total_real_estate_expenses,
        "total_real_estate_gain_loss": total_real_estate_gain_loss,
        "total_real_estate_roi": total_real_estate_roi,

        "total_other_invested": total_other_invested,
        "total_other_value": total_other_value,
        "total_other_income": total_other_income,
        "total_other_expenses": total_other_expenses,
        "total_other_gain_loss": total_other_gain_loss,
        "total_other_roi": total_other_roi,

        "total_sacco_invested": total_sacco_invested,
        "total_sacco_value": total_sacco_value,
        "total_sacco_dividends": total_sacco_dividends,
        "total_sacco_gain_loss": total_sacco_gain_loss,
        "total_sacco_roi": total_sacco_roi,

        "total_current_value": total_current_value,
        "total_pct_gain": total_pct_gain,

        "existing_asset_types": existing_asset_types,
    }


    return render(request, "budget/portfolio_dashboard.html", context)





@login_required
def update_prices_view(request):
    update_prices(request.user)
    messages.success(request, "Prices updated successfully!")
    return redirect("portfolio_dashboard")




@login_required
def add_investment_account(request):
    if request.method == 'POST':
        form = InvestmentAccountForm(request.POST)

        if form.is_valid():
            investment = form.save(commit=False)
            investment.user = request.user

            # Prevent duplicate account names
            duplicate = InvestmentAccount.objects.filter(
                user=request.user,
                name__iexact=investment.name
            ).exclude(pk=investment.pk).exists()
            if duplicate:
                messages.warning(request, f"An investment account named '{investment.name}' already exists. Please choose a different name.")
                return render(request, 'budget/add_investment_account.html', {'form': form})

            manual_price = form.cleaned_data.get("manual_price")
            manual_currency = form.cleaned_data.get("manual_currency") or "KES"

            try:
                investment.full_clean()
                investment.save()

                if manual_price:
                    PriceHistory.objects.create(
                        account=investment,
                        date=date.today(),
                        price=Decimal(manual_price),
                        currency=manual_currency.upper()
                    )

                # 🔹 Ensure "Investments" category exists
                category, _ = ExpenseCategory.objects.get_or_create(
                    name="Investments",
                    user=request.user,
                    defaults={"is_investment_generated": True}  # flag it if you want it excluded from normal expenses
                )

                # 🔹 Create a subcategory for this investment account (so it shows up uniquely in budgets)
                subcategory, _ = ExpenseSubCategory.objects.get_or_create(
                    category=category,
                    name=investment.name,
                    user=request.user,
                )

                # 🔹 Always add a zero-amount expense (so it shows up in budgets)
                Expense.objects.create(
                    user=request.user,
                    category=category,
                    subcategory=subcategory,
                    description=f"Investment account created: {investment.name}",
                    amount=0,
                    date=date.today()
                )

                messages.success(
                    request, 
                    f"'{investment.name}' ({investment.get_asset_type_display()}) successfully added to your portfolio!"
                )
                return redirect('portfolio_dashboard')

            except ValidationError as e:
                error_msg = "; ".join(
                    [f"{field}: {', '.join(msgs)}" for field, msgs in e.message_dict.items()]
                )
                messages.error(request, f"❌ Error saving investment: {error_msg}")

            except Exception as e:
                messages.error(request, f"❌ Could not fetch live data for this ticker: {str(e)}")

        else:
            # Invalid form
            form_errors = "; ".join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()])
            messages.error(request, f"❌ {form_errors or 'Please check your inputs.'}")

        return render(request, 'budget/add_investment_account.html', {'form': form})

    else:
        form = InvestmentAccountForm()
        return render(request, 'budget/add_investment_account.html', {'form': form})


def add_transaction(request, account_id):
    account = get_object_or_404(InvestmentAccount, id=account_id, user=request.user)

    if request.method == "POST":
        form = InvestmentTransactionForm(request.POST, account=account)  # pass account
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.account = account

            # ✅ Amount logic depending on action
            if transaction.action in ["BUY", "SELL"]:
                transaction.amount = (transaction.units or 0) * (transaction.price_per_unit or 0)
            else:
                # Use amount entered by user for dividends, interest, rental income, expenses
                transaction.amount = transaction.amount or 0

            transaction.save()

            # Add success message
            messages.success(
                request,
                f"Transaction for '{account.name}' ({account.get_asset_type_display()}) successfully added!"
            )

            return redirect("portfolio_dashboard")
    else:
        form = InvestmentTransactionForm(account=account)  # also here

    return render(request, "budget/add_transaction.html", {"form": form, "account": account})



@login_required
def delete_investment_account(request, account_id):
    account = get_object_or_404(InvestmentAccount, id=account_id, user=request.user)

    if request.method == "POST":
        account.delete()
        messages.success(request, f"'{account.name}' successfully deleted.")
        return redirect('portfolio_dashboard')

    return redirect('portfolio_dashboard')


@login_required
def update_investment_account(request, account_id):
    account = get_object_or_404(InvestmentAccount, id=account_id, user=request.user)

    if request.method == "POST":
        form = InvestmentAccountForm(request.POST, instance=account)

        manual_price = request.POST.get("manual_price")
        if manual_price in [None, ""]:  
            manual_price = 0
        else:
            manual_price = Decimal(manual_price)


        manual_currency = request.POST.get("manual_currency") or account.currency or "KES"

        if form.is_valid():
            updated_account = form.save(commit=False)
            updated_account.user = request.user

            # Check for duplicate name (case-insensitive) excluding this account
            duplicate = InvestmentAccount.objects.filter(
                user=request.user,
                name__iexact=updated_account.name
            ).exclude(pk=account.pk).exists()

            if duplicate:
                messages.warning(request, f"⚠️ An account named '{updated_account.name}' already exists. Please choose a different name.")
                return render(request, 'budget/add_investment_account.html', {
                    "form": form,
                    "manual_price": manual_price,
                    "manual_currency": manual_currency,
                    "is_update": True,
                    "account": account,
                })

            updated_account.save()

            # Update or create today's PriceHistory if manual price is provided
            if manual_price:
                PriceHistory.objects.update_or_create(
                    account=account,
                    date=date.today(),
                    defaults={
                        "price": Decimal(manual_price),
                        "currency": manual_currency.upper()
                    }
                )

            messages.success(request, "Account updated successfully!")
            return redirect("portfolio_dashboard")

        else:
            # Form invalid
            form_errors = "; ".join([f"{field}: {', '.join(errors)}" for field, errors in form.errors.items()])
            messages.error(request, f"❌ {form_errors or 'Please check your inputs.'}")

    else:
        form = InvestmentAccountForm(instance=account)
        # Populate manual fields
        latest_price_entry = PriceHistory.objects.filter(account=account).order_by("-date").first()
        manual_price = latest_price_entry.price if latest_price_entry else ""
        manual_currency = latest_price_entry.currency if latest_price_entry else ""

    return render(request, 'budget/add_investment_account.html', {
        "form": form,
        "manual_price": manual_price,
        "manual_currency": manual_currency,
        "is_update": True,
        "account": account,
    })


@login_required
def add_wallet_transfer(request):
    if request.method == "POST":
        form = WalletTransferForm(request.POST, user=request.user)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.user = request.user
            transfer.save()
            return redirect("wallet_overview")  # wherever balances are displayed
    else:
        form = WalletTransferForm(user=request.user)
    return render(request, "budget/add_transfer.html", {"form": form})
